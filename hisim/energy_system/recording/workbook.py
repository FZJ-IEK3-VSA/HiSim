"""The two-sheet workbook a person fills in, and where everything in it sits.

The grouping pass needs a human judgement about every difference between a setup's configurations,
and the honest shape of that request is a table: a hundred rows down the left, one column per probe
across, and two empty columns at the right. A spreadsheet is where a person can see the whole thing
at once, sort it, and work down the rows that are not uniform; a YAML file is where the answer
belongs afterwards. This module writes the request; :mod:`hisim.energy_system.recording.workbook_import`
reads the answer back.

``components`` carries one row per component in the union of the probe runs, its class, one cell per
probe holding ``—``, ``=`` or ``≠``, and the two cells the person writes in. ``configurations``
carries one row per probe, the module-configuration fields that produced it, and the switch positions
that column stands for. Both sheets are regenerated from the probe runs whenever they are wanted,
which is why neither is committed.

Regenerating must not throw a decision away, so the writer takes the committed decision as input when
one exists and fills the two right-hand columns from it. That is what makes the round-trip an
identity: probe again, re-import, and the same ``<stem>.grouping.yaml`` comes out, with the rows that
genuinely changed being the only difference.

The setup and the probe list travel in the workbook's own document properties rather than in a third
sheet, so that the file stays exactly the two sheets the requirement describes and a person cannot
accidentally sort a metadata row into the middle of the table.
"""

# clean

from __future__ import annotations

from pathlib import Path
from typing import ClassVar, Dict, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from hisim.energy_system.recording.grouping import Assignment, ConfigurationSelection, Grouping
from hisim.energy_system.recording.matrix import ProbeMatrix
from hisim.energy_system.recording.probes import ProbeList


class WorkbookLayout:
    """Where everything sits in the workbook, written down once for the writer and the reader.

    Two programs read this file — the one that writes it and the one that imports it back — and a
    header spelled differently in the two is the classic way a table tool starts losing a column.
    Keeping the sheet names, the fixed headers and the property encoding as class-scope constants
    means a change to the layout is one edit and both ends follow it.
    """

    #: The sheet holding one row per component of the union of the probe runs.
    COMPONENTS: ClassVar[str] = "components"

    #: The sheet holding one row per probe column.
    CONFIGURATIONS: ClassVar[str] = "configurations"

    #: The fixed columns that come before the probe columns on the components sheet.
    COMPONENT_HEAD: ClassVar[Tuple[str, ...]] = ("component", "class")

    #: The two columns the person writes in, which come after the probe columns.
    COMPONENT_TAIL: ClassVar[Tuple[str, ...]] = ("assignment", "note")

    #: The columns of the configurations sheet, in order.
    CONFIGURATION_HEAD: ClassVar[Tuple[str, ...]] = ("probe", "module_config", "groups", "variants")

    #: How the setup and the probe list are stored in the workbook's document properties.
    PROPERTY_FORMAT: ClassVar[str] = "setup={setup}\nprobes={probes}"

    #: Separator between two switch positions written in one cell.
    POSITION_SEPARATOR: ClassVar[str] = ", "

    #: Separator between a switch's name and its position.
    POSITION_ASSIGNMENT: ClassVar[str] = "="

    #: The spellings a group flag may be written with, mapped to what they mean.
    FLAGS: ClassVar[Dict[str, bool]] = {
        "true": True, "on": True, "yes": True, "1": True,
        "false": False, "off": False, "no": False, "0": False,
    }

    #: Widths of the fixed columns, so that the sheet opens readable rather than as a wall of ####.
    WIDTHS: ClassVar[Dict[str, int]] = {"component": 32, "class": 62, "note": 60, "module_config": 46}

    #: Width every probe column and every switch column gets.
    DEFAULT_WIDTH: ClassVar[int] = 22


class WorkbookWriter:
    """Writes the prefilled workbook from a probe matrix and, when there is one, a prior decision.

    The writer's whole job is to make the request as small as possible: everything observation can
    answer is already in the cell, and the only empty cells are the ones that need a judgement. A
    row that is ``=`` in every column needs none and is written anyway, because a person deciding
    about the meter wants to see the meter next to the rest of the system rather than in a filtered
    list that hides what it is being compared with.

    Prefilling the two right-hand columns from an existing decision is what makes re-probing safe.
    A setup that grew a component changes one row of the table; without the prefill it would throw
    the other ninety-nine decisions away and ask for all of them again.
    """

    @classmethod
    def write(
        cls,
        path: Path,
        matrix: ProbeMatrix,
        probe_list: ProbeList,
        grouping: Optional[Grouping] = None,
    ) -> Path:
        """Writes the whole workbook.

        Args:
            path: Where the file goes; parent directories are created.
            matrix: The three-state table to prefill from.
            probe_list: The probes, for the column order and the fields that produced each column.
            grouping: A decision already committed for this setup, whose assignments and switch
                positions are carried over; ``None`` when there is none yet.

        Returns:
            The path written.
        """
        workbook = Workbook()
        workbook.properties.title = f"{Path(probe_list.setup).stem} grouping"
        workbook.properties.description = WorkbookLayout.PROPERTY_FORMAT.format(
            setup=probe_list.setup, probes=probe_list.origin
        )
        components = workbook.active
        components.title = WorkbookLayout.COMPONENTS
        cls._components(components, matrix, grouping)
        cls._configurations(
            workbook.create_sheet(WorkbookLayout.CONFIGURATIONS), matrix, probe_list, grouping
        )
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path

    @classmethod
    def _components(cls, sheet: Worksheet, matrix: ProbeMatrix, grouping: Optional[Grouping]) -> None:
        """Fills the components sheet.

        Args:
            sheet: The sheet to write.
            matrix: The three-state table.
            grouping: The prior decision, or ``None``.
        """
        header = (*WorkbookLayout.COMPONENT_HEAD, *matrix.columns, *WorkbookLayout.COMPONENT_TAIL)
        cls._header(sheet, header)
        for row in matrix.rows:
            decision = grouping.assignment(row.component) if grouping is not None else Assignment(row.component)
            sheet.append(
                [
                    row.component,
                    row.class_path,
                    *(row.states[column].value for column in matrix.columns),
                    decision.text,
                    decision.note,
                ]
            )
        cls._widths(sheet, header)

    @classmethod
    def _configurations(
        cls, sheet: Worksheet, matrix: ProbeMatrix, probe_list: ProbeList, grouping: Optional[Grouping]
    ) -> None:
        """Fills the configurations sheet.

        Args:
            sheet: The sheet to write.
            matrix: The table, for the columns that were really recorded.
            probe_list: The probes, for the fields that produced each column.
            grouping: The prior decision, or ``None``.
        """
        cls._header(sheet, WorkbookLayout.CONFIGURATION_HEAD)
        for column in matrix.columns:
            selection = (
                grouping.selection(column) if grouping is not None else ConfigurationSelection(column=column)
            )
            sheet.append(
                [
                    column,
                    WorkbookLayout.POSITION_SEPARATOR.join(probe_list.probe(column).fields()),
                    cls._positions({name: str(flag).lower() for name, flag in selection.groups.items()}),
                    cls._positions(dict(selection.variants)),
                ]
            )
        cls._widths(sheet, WorkbookLayout.CONFIGURATION_HEAD)

    @classmethod
    def _positions(cls, positions: Dict[str, str]) -> str:
        """Renders one cell holding several ``name=position`` pairs.

        Args:
            positions: The switch positions.

        Returns:
            The rendered cell, empty when there are none.
        """
        return WorkbookLayout.POSITION_SEPARATOR.join(
            f"{name}{WorkbookLayout.POSITION_ASSIGNMENT}{value}" for name, value in positions.items()
        )

    @classmethod
    def _header(cls, sheet: Worksheet, header: Sequence[str]) -> None:
        """Writes and freezes one sheet's header row.

        Args:
            sheet: The sheet.
            header: The column titles.
        """
        sheet.append(list(header))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        sheet.freeze_panes = "A2"

    @classmethod
    def _widths(cls, sheet: Worksheet, header: Sequence[str]) -> None:
        """Gives every column a width wide enough to read without resizing.

        Args:
            sheet: The sheet.
            header: The column titles, which decide which width each column gets.
        """
        for index, title in enumerate(header, start=1):
            width = WorkbookLayout.WIDTHS.get(title, WorkbookLayout.DEFAULT_WIDTH)
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width


def write_workbook(
    path: Path, matrix: ProbeMatrix, probe_list: ProbeList, grouping: Optional[Grouping] = None
) -> Path:
    """Writes the prefilled grouping workbook for one setup.

    Args:
        path: Where the file goes.
        matrix: The three-state table.
        probe_list: The probes the table's columns come from.
        grouping: A decision already committed for this setup, carried over into the two columns
            the person writes in; ``None`` when the setup has none yet.

    Returns:
        The path written.
    """
    return WorkbookWriter.write(path, matrix, probe_list, grouping)
