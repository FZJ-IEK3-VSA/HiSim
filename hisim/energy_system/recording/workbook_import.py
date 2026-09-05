"""Reading a filled-in grouping workbook back into the decision it states.

The importer takes the two right-hand columns of the ``components`` sheet and the two switch columns
of the ``configurations`` sheet as the decision. The three-state cells are read for one early check
and nothing else: the rule that a component which is not the same in every column must carry a
decision is applied here, against the cells as written, because the person has the file open and a
refusal naming the row and the columns is one edit away from being fixed rather than four probe
runs away. The cells are the tool's own output, but a person *can* type over one — which is why the
check that counts is the second pass's, made against the real recordings, where a typed-over cell
has changed nothing.

Every refusal names the sheet and the row, because a spreadsheet with a hundred rows is the one place
where "something is wrong with the file" is a useless message.
"""

# clean

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hisim.energy_system.errors import EnergySystemErrorId, EnergySystemRecordingError
from hisim.energy_system.recording.grouping import Assignment, ConfigurationSelection, Grouping
from hisim.energy_system.recording.matrix import CellState
from hisim.energy_system.recording.workbook import WorkbookLayout


class WorkbookReader:
    """Reads a filled-in workbook back into the decision it states.

    The reader takes the two right-hand columns of the components sheet and the two switch columns
    of the configurations sheet as the decision, and reads the three-state cells only for the early
    undecided-row check. A typed-over state cell can weaken that early check and nothing else: the
    authoritative one is made by the second pass against the real recordings.

    Every refusal names the sheet and the row, because a spreadsheet with a hundred rows is the one
    place where "something is wrong with the file" is a useless message.
    """

    @classmethod
    def read(cls, path: Path) -> Grouping:
        """Reads one workbook.

        Args:
            path: The ``.xlsx`` file.

        Returns:
            The decision it states, with the setup and probe list taken from the workbook's
            properties and falling back to the conventional names beside it.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when a sheet is missing or a switch cell cannot
                be read, and ``EF-R6`` when an assignment cell is not one of the four forms.
        """
        workbook = load_workbook(path, read_only=False, data_only=True)
        for sheet_name in (WorkbookLayout.COMPONENTS, WorkbookLayout.CONFIGURATIONS):
            if sheet_name not in workbook.sheetnames:
                raise cls._error(path, sheet_name, f"the workbook has no '{sheet_name}' sheet.")
        setup, probes = cls._properties(workbook.properties.description, path)
        return Grouping(
            setup=setup,
            probes=probes,
            assignments=cls._assignments(workbook[WorkbookLayout.COMPONENTS], path),
            configurations=cls._configurations(workbook[WorkbookLayout.CONFIGURATIONS], path),
            origin=str(path),
        )

    @classmethod
    def _properties(cls, description: Optional[str], path: Path) -> Tuple[str, str]:
        """Recovers the setup and the probe list the workbook was written for.

        The writer always records both in the workbook's description property, so their absence
        means the provenance was lost — a spreadsheet tool that strips document properties on save
        is the usual culprit. Guessing the paths from the filename would commit a decision whose
        setup/probes fields nobody stated, and the wrong guess would only surface later, loudly, at
        ``record --grouping`` — after the decision was committed. Lost provenance is refused here
        instead.

        Args:
            description: The workbook's description property, or ``None`` when it carries none.
            path: The workbook, for the message.

        Returns:
            The setup path and the probe list path.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when either is missing from the properties.
        """
        found = dict(
            line.split("=", 1) for line in (description or "").splitlines() if "=" in line
        )
        for key in ("setup", "probes"):
            if not found.get(key, "").strip():
                raise cls._error(
                    path,
                    "document properties",
                    f"the workbook does not say which {key} it was written for; its description "
                    "property was probably stripped by the tool that saved it. Re-run 'grouping "
                    "probe' to regenerate the workbook and carry the filled-in decision over.",
                )
        return found["setup"].strip(), found["probes"].strip()

    @classmethod
    def _assignments(cls, sheet: Worksheet, path: Path) -> Tuple[Assignment, ...]:
        """Reads the two right-hand columns of the components sheet.

        Args:
            sheet: The components sheet.
            path: The workbook, for the message.

        Returns:
            One assignment per row that says anything, in sheet order.

        Raises:
            EnergySystemRecordingError: ``EF-R6`` for an unreadable assignment cell and ``EF-R7``
                when the sheet has no header at all.
        """
        header = cls._header(sheet, path)
        for title in (*WorkbookLayout.COMPONENT_HEAD, *WorkbookLayout.COMPONENT_TAIL):
            if title not in header:
                raise cls._error(path, WorkbookLayout.COMPONENTS, f"the sheet has no '{title}' column.")
        columns = cls._probe_columns(header)
        assignments: List[Assignment] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            component = cls._text(values, header, "component")
            if not component:
                continue
            written = Assignment.parse(
                component,
                cls._text(values, header, "assignment"),
                cls._text(values, header, "note"),
                str(path),
            )
            cls._check_decided(component, written, cls._states(values, header, columns), path)
            if written.text or written.note:
                assignments.append(written)
        return tuple(assignments)

    @classmethod
    def _probe_columns(cls, header: Dict[str, int]) -> Tuple[str, ...]:
        """The probe columns of the components sheet, which are whatever sits between the fixed ones.

        Reading them from the sheet rather than from the probe list is deliberate: the importer
        works on the workbook alone, and a person who added a column by hand should see the rule
        applied to it rather than have it silently ignored.

        Args:
            header: The column index of every title.

        Returns:
            The probe column titles, in sheet order.
        """
        fixed = (*WorkbookLayout.COMPONENT_HEAD, *WorkbookLayout.COMPONENT_TAIL)
        return tuple(title for title in header if title not in fixed)

    @classmethod
    def _states(
        cls, values: Sequence[Any], header: Dict[str, int], columns: Tuple[str, ...]
    ) -> Dict[str, Optional[CellState]]:
        """Reads one row's three-state cells back.

        Args:
            values: The row's values.
            header: The column index of every title.
            columns: The probe columns.

        Returns:
            The state of every probe column, ``None`` where the cell holds something else.
        """
        return {column: CellState.of(cls._text(values, header, column)) for column in columns}

    @classmethod
    def _check_decided(
        cls,
        component: str,
        assignment: Assignment,
        states: Dict[str, Optional[CellState]],
        path: Path,
    ) -> None:
        """Refuses a row that is not the same everywhere and carries no decision.

        The same rule is checked again by the second pass against the real recordings, and it is
        checked here as well because here is where it can be acted on: the person has the workbook
        open, and a refusal naming the row and the columns is one edit away from being fixed.

        Args:
            component: The row's component.
            assignment: What the row says.
            states: The row's three-state cells.
            path: The workbook, for the message.

        Raises:
            EnergySystemRecordingError: ``EF-R6`` naming the row and the columns that are not ``=``.
        """
        unequal = tuple(
            column for column, state in states.items() if state is not CellState.IDENTICAL
        )
        if not unequal or assignment.text:
            return
        raise EnergySystemRecordingError(
            EnergySystemErrorId.GROUPING_UNASSIGNED_DIFFERENCE,
            f"{path}:{WorkbookLayout.COMPONENTS}.{component}",
            f"'{component}' is not the same in {', '.join(unequal)} and its assignment cell is empty.",
            alternatives=("override", "group:<name>", "variant:<name>", "variant:<name>/<option>"),
            alternatives_label="assignments",
            remedy=(
                "A component cannot both stay ungrouped and disagree with itself between two "
                "configurations. Say what the difference means in the row's assignment cell."
            ),
        )

    @classmethod
    def _configurations(cls, sheet: Worksheet, path: Path) -> Tuple[ConfigurationSelection, ...]:
        """Reads the switch positions of the configurations sheet.

        Args:
            sheet: The configurations sheet.
            path: The workbook, for the message.

        Returns:
            One selection per probe row, in sheet order.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when a column is missing or a cell is malformed.
        """
        header = cls._header(sheet, path)
        for title in WorkbookLayout.CONFIGURATION_HEAD:
            if title not in header:
                raise cls._error(path, WorkbookLayout.CONFIGURATIONS, f"the sheet has no '{title}' column.")
        selections: List[ConfigurationSelection] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            column = cls._text(values, header, "probe")
            if not column:
                continue
            selections.append(
                ConfigurationSelection(
                    column=column,
                    groups=cls._flags(cls._text(values, header, "groups"), column, path),
                    variants=cls._pairs(cls._text(values, header, "variants"), column, path),
                )
            )
        return tuple(selections)

    @classmethod
    def _flags(cls, cell: str, column: str, path: Path) -> Dict[str, bool]:
        """Reads a cell of ``name=on`` group flags.

        Args:
            cell: The cell's text.
            column: The probe column the row is about, for the message.
            path: The workbook, for the message.

        Returns:
            The flags.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when a position is not a readable flag.
        """
        flags: Dict[str, bool] = {}
        for name, value in cls._pairs(cell, column, path).items():
            if value.lower() not in WorkbookLayout.FLAGS:
                raise cls._error(
                    path,
                    f"{WorkbookLayout.CONFIGURATIONS}.{column}.groups",
                    f"'{value}' is not a position a group flag can be in.",
                )
            flags[name] = WorkbookLayout.FLAGS[value.lower()]
        return flags

    @classmethod
    def _pairs(cls, cell: str, column: str, path: Path) -> Dict[str, str]:
        """Reads a cell of comma-separated ``name=value`` pairs.

        Args:
            cell: The cell's text.
            column: The probe column the row is about, for the message.
            path: The workbook, for the message.

        Returns:
            The pairs, in written order.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when an item carries no ``=`` at all.
        """
        pairs: Dict[str, str] = {}
        for item in (piece.strip() for piece in cell.split(",")):
            if not item:
                continue
            name, separator, value = item.partition(WorkbookLayout.POSITION_ASSIGNMENT)
            if not separator or not name.strip() or not value.strip():
                raise cls._error(
                    path,
                    f"{WorkbookLayout.CONFIGURATIONS}.{column}",
                    f"'{item}' is not a 'name=position' pair.",
                )
            pairs[name.strip()] = value.strip()
        return pairs

    @classmethod
    def _header(cls, sheet: Worksheet, path: Path) -> Dict[str, int]:
        """Indexes one sheet's header row by column title.

        Args:
            sheet: The sheet.
            path: The workbook, for the message.

        Returns:
            The column index of every title, zero-based.

        Raises:
            EnergySystemRecordingError: ``EF-R7`` when the sheet is empty.
        """
        rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        titles = next(iter(rows), None)
        if titles is None:
            raise cls._error(path, str(sheet.title), "the sheet has no header row.")
        return {str(title): index for index, title in enumerate(titles) if title is not None}

    @classmethod
    def _text(cls, values: Sequence[Any], header: Dict[str, int], title: str) -> str:
        """Reads one cell of one row as trimmed text.

        Args:
            values: The row's values.
            header: The column index of every title.
            title: The column wanted.

        Returns:
            The cell's text, empty when the cell is empty or the row is short.
        """
        index = header.get(title)
        if index is None or index >= len(values) or values[index] is None:
            return ""
        return str(values[index]).strip()

    @classmethod
    def _error(cls, path: Path, location: str, problem: str) -> EnergySystemRecordingError:
        """Builds one workbook rejection; the caller raises it.

        Args:
            path: The workbook.
            location: The sheet, and the row where one is known.
            problem: One sentence naming what is wrong.

        Returns:
            The exception.
        """
        return EnergySystemRecordingError(
            EnergySystemErrorId.GROUPING_UNKNOWN_OPTION, f"{path}:{location}", problem
        )


def read_workbook(path: Path) -> Grouping:
    """Reads a filled-in grouping workbook back into the decision it states.

    Args:
        path: The ``.xlsx`` file.

    Returns:
        The decision.

    Raises:
        EnergySystemRecordingError: ``EF-R6`` for an unreadable or undecided row, ``EF-R7`` for a
            malformed sheet.
    """
    return WorkbookReader.read(path)
