import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import openpyxl
import win32com.client


# Pfad zum Hauptordner, in dem sich die Unterordner mit Excel-Files befinden
main_folder = 'C://Users//Standard//Desktop//hisim//results//OnlyBattery//'
excel_filename = 'C://Users//Standard//Desktop//hisim//results//OnlyBattery//Gesamtergebnis.xlsx'

#main_folder = 'C://Users\Standard//Desktop//hisim//C4LResults//Auswertung//'
#excel_filename = 'C://Users//Standard//Desktop//hisim//C4LResults//Auswertung//Gesamtergebnis.xlsx'


def startexcelsavefile(excel_file_path):

    excel = win32com.client.Dispatch("Excel.Application")
    
    # Arbeitsmappe öffnen
    excel_file_path = excel_file_path.replace('//', '\\')
    wb = excel.Workbooks.Open(excel_file_path)

    # Hier könnten zusätzliche Aktionen in Excel ausgeführt werden

    # Arbeitsmappe speichern
    wb.Save()

    # Arbeitsmappe schließen
    wb.Close()

    # Excel-Anwendung beenden
    excel.Quit()

# Funktion, um den gewünschten Wert aus dem Excel-File zu extrahieren
def extract_value_from_excel(file_path):


    # Dies könnte mithilfe von pandas oder openpyxl geschehen, abhängig von der Struktur der Daten
    # Hier ist ein Beispiel, wie du eine Spalte 'Wert' aus dem Excel-File liest:
    df = pd.read_excel(file_path)
    
    # Lade die Excel-Datei
    workbook = load_workbook(file_path, data_only=True)

    #Wirtschaftliche Betrachtung
    worksheet = workbook['Eingaben']
    jaehrlichesbetriebsergebnis_inEuro = worksheet['C126'].value
    Vjaehrlichesbetriebsergebnis_inEuro = worksheet['D126'].value


    worksheet = workbook['Jahresreihe+Vergleichsjahresr']
    kapitalwertdifferenz_nach20jahren_inEuro_SOFC_Versus_NURPV = worksheet['L23'].value
    kapitalwert_nach20jahren_inEuro = worksheet['I23'].value

    #Autarkiequote Strom
    worksheet = workbook['Ergebnisse-Energie']
    
    netzbezugsanteil_elektr_Prozent= worksheet['B108'].value
    eigenerzeugungsanteil_elektr_Prozent = worksheet['B109'].value
    gesamtstrombedarf_elektr_MWh = worksheet['B113'].value
    eigenerzeugungsanteil_elektr_MWh = worksheet['B112'].value
    netzbezugsanteil_elektr_MWh= worksheet['B111'].value

    Vnetzbezugsanteil_elektr_Prozent= worksheet['C108'].value
    Veigenerzeugungsanteil_elektr_Prozent = worksheet['C109'].value
    Vgesamtstrombedarf_elektr_MWh = worksheet['C113'].value
    Veigenerzeugungsanteil_elektr_MWh = worksheet['C112'].value
    Vnetzbezugsanteil_elektr_MWh= worksheet['C111'].value
    
    #Eigenverbrauchsquote Strom
    gesamterzeugungPV_MWh = worksheet['B102'].value
    netzeinspeisungsquote_elektr_Prozent = worksheet['B97'].value
    eigenverbrauchsquote_elektr_Prozent = worksheet['B98'].value
    netzeinspeisungsquote_elektr_MWh = worksheet['B100'].value    
    eigenverbrauchsquote_elektr_MWh = worksheet['B101'].value

    VgesamterzeugungPV_MWh = worksheet['C102'].value
    Vnetzeinspeisungsquote_elektr_Prozent = worksheet['C97'].value
    Veigenverbrauchsquote_elektr_Prozent = worksheet['C98'].value
    Vnetzeinspeisungsquote_elektr_MWh = worksheet['C100'].value    
    Veigenverbrauchsquote_elektr_MWh = worksheet['C101'].value

    #Wärmeenergie
    gesamtwaermebedarf_MWh = worksheet['B123'].value 
    waermebedarfsanteil_MWh = worksheet['B121'].value
    eigenerzeugungsanteil_MWh = worksheet['B122'].value    
    waermebedarfsanteil_Prozent = worksheet['B118'].value
    eigenerzeugungsanteil_Prozent = worksheet['B119'].value 

    Vgesamtwaermebedarf_MWh = worksheet['C123'].value 
    Vwaermebedarfsanteil_MWh = worksheet['C121'].value
    Veigenerzeugungsanteil_MWh = worksheet['C122'].value  
    Vwaermebedarfsanteil_Prozent = worksheet['C118'].value
    Veigenerzeugungsanteil_Prozent = worksheet['C119'].value 

    #Anlagendaten
    worksheet = workbook['Anlagendaten']
    batterycapacity_kWh = worksheet['C2'].value 
    batterinverwertepower_kW = worksheet['C3'].value/1000
    fuelcellpower_kW = worksheet['C4'].value/1000
    electrolyzer_kW = worksheet['C11'].value/1000
    
    worksheet = workbook['Eingaben']
    h2storage_capacity_kg = worksheet['C44'].value
    h2storage_kubikmeter = worksheet['C45'].value
    h2storage_pressure = worksheet['B45'].value

    return kapitalwert_nach20jahren_inEuro, kapitalwertdifferenz_nach20jahren_inEuro_SOFC_Versus_NURPV, jaehrlichesbetriebsergebnis_inEuro,  gesamtstrombedarf_elektr_MWh, eigenerzeugungsanteil_elektr_MWh, netzbezugsanteil_elektr_MWh, netzbezugsanteil_elektr_Prozent, eigenerzeugungsanteil_elektr_Prozent, gesamterzeugungPV_MWh, eigenverbrauchsquote_elektr_MWh, netzeinspeisungsquote_elektr_MWh, netzeinspeisungsquote_elektr_Prozent, eigenverbrauchsquote_elektr_Prozent, gesamtwaermebedarf_MWh, eigenerzeugungsanteil_MWh, waermebedarfsanteil_MWh, eigenerzeugungsanteil_Prozent, waermebedarfsanteil_Prozent,\
        Vjaehrlichesbetriebsergebnis_inEuro,  Vgesamtstrombedarf_elektr_MWh, Veigenerzeugungsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_Prozent, Veigenerzeugungsanteil_elektr_Prozent, VgesamterzeugungPV_MWh, Veigenverbrauchsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_Prozent, Veigenverbrauchsquote_elektr_Prozent, Vgesamtwaermebedarf_MWh, Veigenerzeugungsanteil_MWh, Vwaermebedarfsanteil_MWh, Veigenerzeugungsanteil_Prozent, Vwaermebedarfsanteil_Prozent, \
        batterycapacity_kWh, batterinverwertepower_kW, fuelcellpower_kW, electrolyzer_kW, h2storage_capacity_kg, h2storage_kubikmeter, h2storage_pressure 


# Liste für die gesammelten Werte und Pfade
values_and_paths = []

# Durchlaufe alle Unterordner im Hauptordner
for foldername in os.listdir(main_folder):
    print(foldername)
    folder_path = os.path.join(main_folder, foldername)
    if os.path.isdir(folder_path):
        # Liste für die Excel-Files im aktuellen Unterordner
        excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]
        # Durchlaufe die Excel-Files im aktuellen Unterordner
        for excel_file in excel_files:
            
            file_path = os.path.join(folder_path, excel_file)
            # Hier prüfen wir, ob die ersten 10 Buchstaben des Dateinamens übereinstimmen
            if excel_file[2:18] == '_Oek_Assessment_' or excel_file[3:19] == '_Oek_Assessment_' or excel_file[4:20] == '_Oek_Assessment_' or excel_file[5:21] == '_Oek_Assessment_' or excel_file[6:22] == '_Oek_Assessment_':
                print(foldername)
                startexcelsavefile(file_path)
                # Extrahiere den Wert aus dem Excel-File
                kapitalwert_nach20jahren_inEuro, kapitalwertdifferenz_nach20jahren_inEuro_SOFC_Versus_NURPV, jaehrlichesbetriebsergebnis_inEuro,  gesamtstrombedarf_elektr_MWh, eigenerzeugungsanteil_elektr_MWh, netzbezugsanteil_elektr_MWh, netzbezugsanteil_elektr_Prozent, eigenerzeugungsanteil_elektr_Prozent, gesamterzeugungPV_MWh, eigenverbrauchsquote_elektr_MWh, netzeinspeisungsquote_elektr_MWh, netzeinspeisungsquote_elektr_Prozent, eigenverbrauchsquote_elektr_Prozent, gesamtwaermebedarf_MWh, eigenerzeugungsanteil_MWh, waermebedarfsanteil_MWh, eigenerzeugungsanteil_Prozent, waermebedarfsanteil_Prozent,\
                    Vjaehrlichesbetriebsergebnis_inEuro,  Vgesamtstrombedarf_elektr_MWh, Veigenerzeugungsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_Prozent, Veigenerzeugungsanteil_elektr_Prozent, VgesamterzeugungPV_MWh, Veigenverbrauchsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_Prozent, Veigenverbrauchsquote_elektr_Prozent, Vgesamtwaermebedarf_MWh, Veigenerzeugungsanteil_MWh, Vwaermebedarfsanteil_MWh, Veigenerzeugungsanteil_Prozent, Vwaermebedarfsanteil_Prozent, \
                    batterycapacity_kWh, batterinverwertepower_kW, fuelcellpower_kW, electrolyzer_kW, h2storage_capacity_kg,  h2storage_kubikmeter, h2storage_pressure, \
                        = extract_value_from_excel(file_path)

                

                # Füge den Wert und den Dateipfad zur Liste hinzu

                values_and_paths.append((kapitalwert_nach20jahren_inEuro, kapitalwertdifferenz_nach20jahren_inEuro_SOFC_Versus_NURPV, jaehrlichesbetriebsergebnis_inEuro,  gesamtstrombedarf_elektr_MWh, eigenerzeugungsanteil_elektr_MWh, netzbezugsanteil_elektr_MWh, netzbezugsanteil_elektr_Prozent, eigenerzeugungsanteil_elektr_Prozent, gesamterzeugungPV_MWh, eigenverbrauchsquote_elektr_MWh, netzeinspeisungsquote_elektr_MWh, netzeinspeisungsquote_elektr_Prozent, eigenverbrauchsquote_elektr_Prozent, gesamtwaermebedarf_MWh, eigenerzeugungsanteil_MWh, waermebedarfsanteil_MWh, eigenerzeugungsanteil_Prozent, waermebedarfsanteil_Prozent,\
                                         file_path, foldername, \
                                             Vjaehrlichesbetriebsergebnis_inEuro,  Vgesamtstrombedarf_elektr_MWh, Veigenerzeugungsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_MWh, Vnetzbezugsanteil_elektr_Prozent, Veigenerzeugungsanteil_elektr_Prozent, VgesamterzeugungPV_MWh, Veigenverbrauchsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_MWh, Vnetzeinspeisungsquote_elektr_Prozent, Veigenverbrauchsquote_elektr_Prozent, Vgesamtwaermebedarf_MWh, Veigenerzeugungsanteil_MWh, Vwaermebedarfsanteil_MWh, Veigenerzeugungsanteil_Prozent, Vwaermebedarfsanteil_Prozent, \
                                                batterycapacity_kWh, batterinverwertepower_kW, fuelcellpower_kW, electrolyzer_kW, h2storage_capacity_kg, h2storage_kubikmeter, h2storage_pressure ))
                
                headers = ['kapitalwert_nach20jahren_inEuro','kapitalwertdifferenz_nach20jahren_inEuro_SOFC_Versus_NURPV', 'jaehrlichesbetriebsergebnis_inEuro',  'gesamtstrombedarf_elektr_MWh', 'eigenerzeugungsanteil_elektr_MWh', 'netzbezugsanteil_elektr_MWh', 'netzbezugsanteil_elektr_Prozent', 'eigenerzeugungsanteil_elektr_Prozent', 'gesamterzeugungPV_MWh', 'eigenverbrauchsquote_elektr_MWh', 'netzeinspeisungsquote_elektr_MWh', 'netzeinspeisungsquote_elektr_Prozent', 'eigenverbrauchsquote_elektr_Prozent', 'gesamtwaermebedarf_MWh', 'eigenerzeugungsanteil_MWh', 'waermebedarfsanteil_MWh', 'eigenerzeugungsanteil_Prozent', 'waermebedarfsanteil_Prozent',\
                           'file_path', 'foldername', \
                           'Vjaehrlichesbetriebsergebnis_inEuro',  'Vgesamtstrombedarf_elektr_MWh', 'Veigenerzeugungsanteil_elektr_MWh', 'Vnetzbezugsanteil_elektr_MWh', 'Vnetzbezugsanteil_elektr_Prozent', 'Veigenerzeugungsanteil_elektr_Prozent', 'VgesamterzeugungPV_MWh', 'Veigenverbrauchsquote_elektr_MWh', 'Vnetzeinspeisungsquote_elektr_MWh', 'Vnetzeinspeisungsquote_elektr_Prozent', 'Veigenverbrauchsquote_elektr_Prozent', 'Vgesamtwaermebedarf_MWh', 'Veigenerzeugungsanteil_MWh', 'Vwaermebedarfsanteil_MWh', 'Veigenerzeugungsanteil_Prozent', 'Vwaermebedarfsanteil_Prozent',\
                            'batterycapacity_kWh', 'batterinverwertepower_kW', 'fuelcellpower_kW', 'electrolyzer_kW', 'h2storage_capacity_kg', 'h2storage_Kubikmeter', 'h2storage_Pressure_bar']

# Sortiere die Werte nach ihrer Größe
sorted_values_and_paths = sorted(values_and_paths, key=lambda x: x[0], reverse=True)


# Erstelle eine neue Excel-Arbeitsmappe
#workbook = Workbook()
# Erstes Tabellenblatt für unsortierte Daten
#sheet_unsorted = workbook.active
# 

workbook = openpyxl.load_workbook(excel_filename)
sheet_unsorted = workbook['Unsortierte Daten']

#sheet_unsorted.title = "Unsortierte Daten"
for row in sheet_unsorted.iter_rows():
    for cell in row:
        cell.value = None


# Füge Überschriften in das erste Tabellenblatt ein
for col_index, header in enumerate(headers, start=1):
    sheet_unsorted.cell(row=1, column=col_index, value=header)

# Füge die unsortierten Daten in das erste Tabellenblatt ein
for row_index, row_data in enumerate(values_and_paths, start=2):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet_unsorted.cell(row=row_index, column=col_index, value=cell_value)


# Zweites Tabellenblatt für sortierte Daten
#sheet_sorted = workbook.create_sheet(title="SortiertKapitalwert")


sheet_sorted  = workbook['SortiertKapitalwert']

#sheet_unsorted.title = "Unsortierte Daten"
for row in sheet_sorted .iter_rows():
    for cell in row:
        cell.value = None


# Füge Überschriften in das zweite Tabellenblatt ein
for col_index, header in enumerate(headers, start=1):
    sheet_sorted.cell(row=1, column=col_index, value=header)

# Füge die sortierten Daten in das zweite Tabellenblatt ein
for row_index, row_data in enumerate(sorted_values_and_paths, start=2):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet_sorted.cell(row=row_index, column=col_index, value=cell_value)


sorted_values_and_paths_two = sorted(values_and_paths, key=lambda x: x[5], reverse=False)

sheet_sorted_netzbezug  = workbook['SortiertNetzbezug']

#sheet_unsorted.title = "Unsortierte Daten"
for row in sheet_sorted_netzbezug .iter_rows():
    for cell in row:
        cell.value = None


# Füge Überschriften in das zweite Tabellenblatt ein
for col_index, header in enumerate(headers, start=1):
    sheet_sorted_netzbezug.cell(row=1, column=col_index, value=header)

# Füge die sortierten Daten in das zweite Tabellenblatt ein
for row_index, row_data in enumerate(sorted_values_and_paths_two, start=2):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet_sorted_netzbezug.cell(row=row_index, column=col_index, value=cell_value)







# Speichere die Excel-Datei

workbook.save(excel_filename)

print("Daten wurden in '{}' gespeichert.".format(excel_filename))



