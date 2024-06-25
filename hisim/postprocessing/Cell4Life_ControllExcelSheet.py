import csv
import os
import json
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import NamedStyle
from hisim.result_path_provider import ResultPathProviderSingleton

def ControllSheetExcel(excelfilepathresults3, input_variablen):


    # #----Save all Input Data in a .txt ----
    # path0 = ResultPathProviderSingleton().get_result_directory_name() 
    # path = path0 + "//"
    # path2 = os.path.join(path,'ControllFileData.xlsx')
    # del path0

    #----Save all Input Data in a .txt ----
    path = ResultPathProviderSingleton().get_result_directory_name() 
    workbook = openpyxl.load_workbook(excelfilepathresults3)
    worksheet = workbook['Modellergebnis']
    #----------------------------------------------   
  
    #Add Input Data for calcuation within "Gesamtmodell"
    #Load Original Energy-Input Data
    model_energy_input_data = []
    EnergyInputData2 = []
    EnergyInputData3 = []
    EnergyInputData4 = []
    EnergyInputData5 = []
    EnergyInputData6 = []
    EnergyInputData7 = []
    EnergyInputData8 = []
    EnergyInputData9 = []
    EnergyInputData10 = []
    EnergyInputData11 = []
    EnergyInputData12 = []
    EnergyInputData13 = []
    EnergyInputData14 = []
    EnergyInputData15 = []

    # Load data from "-_Current" (column 1 & 2) and add the collected data to list
    csv_datei1 = os.path.join(path, 'CSV_PVComponent.csv')
    with open(csv_datei1, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        izaehler = 0
        iii = 1

        for row in csvreader:
            if izaehler == 0:
                monat = 'Monat'
            else:
                datum = row[0]
                datum_obj = datetime.strptime(datum, '%Y-%m-%d %H:%M:%S')
                monat = datum_obj.month
            
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 

            model_energy_input_data.append([row[0], monat, row[1]])
            izaehler += 1 
            iii += 1
    del iii

    # Load data from "Fuel Cell Component" (column 2) and add the collected data to list
    csv_datei2 = os.path.join(path, 'ElectricityOutput_CHP_w2.csv')
    with open(csv_datei2, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            
            EnergyInputData2.append(row[1])
            iii+=1
    del iii

    # Load data from "Ac Battery Component" (column 2) and add the collected data to list
    csv_datei3 = os.path.join(path, 'AcBatteryPower_Battery_w1.csv')
    with open(csv_datei3, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData3.append(row[1])
            iii +=1
    del iii

    # Load data from "Ac Battery Component" (column 2) and add the collected data to list
    csv_datei4 = os.path.join(path, 'StateOfCharge_Battery_w1.csv')
    with open(csv_datei4, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData4.append(row[1])
            iii+=1
    del iii


    # Load data from "-Current needed" (column 2) and add the collected data to list
    csv_datei5= os.path.join(path, '-_Current.csv')
    with open(csv_datei5, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData5.append(row[1])
            iii+=1
    del iii
            


    # Load data from "Electrolyzher Component" (column 2) and add the collected data to list
    #csv_datei6 = os.path.join(path, 'ElectricityConsumptionElectrolyzer_StaticElectrolyzer.csv')
    csv_datei6 = os.path.join(path, 'ElectricityConsumptionElectrolyzer_C4LElectrolyzer.csv')
    with open(csv_datei6, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData6.append(row[1])
            iii+=1
    del iii

    # Load data from "H2 Storage Component" (column 2) and add the collected data to list
    csv_datei7 = os.path.join(path, 'StorageElectricityConsumption_HydrogenStorage_w999.csv')
    with open(csv_datei7, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData7.append(row[1])
            iii+=1
    del iii

    # Load data from "to or from grid" (column 2) and add the collected data to list
    csv_datei8 = os.path.join(path, 'ElectricityToOrFromGrid_Elect_Controller.csv')
    with open(csv_datei8, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData8.append(row[1])
            iii += 1
    del iii
    
    # Load data from "from CHP to house" (column 2) and add the collected data to list
    csv_datei9 = os.path.join(path, 'QuantitiyShare_electricity_from_CHP_to_house_Elect_Controller.csv')
    with open(csv_datei9, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData9.append(row[1])
            iii+=1
    del iii

    # Load data from "from Battery to house" (column 2) and add the collected data to list
    csv_datei10 = os.path.join(path, 'QuantitiyShare_electricity_from_Battery_to_house_Elect_Controller.csv')
    with open(csv_datei10, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData10.append(row[1])
            iii+=1
    del iii


    # Load data from "PV to grid" (column 2) and add the collected data to list
    csv_datei11 = os.path.join(path, 'QuantitiyShare_PV_to_grid_Elect_Controller.csv')
    with open(csv_datei11, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1])
                
            EnergyInputData11.append(row[1])
            iii +=1
    del iii

    # Load data from "Electricity which needs Fuel Cell in Standby" (column 2) and add the collected data to list
    csv_datei12 = os.path.join(path, 'FuelCellElectricityInputStandby_CHP_w2.csv')
    with open(csv_datei12, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1])

            EnergyInputData12.append(row[1])
            iii+=1
    del iii


    # Load data from "the amount of energy which is delivered from the battery to the CHP/Fuel Cell in Standbymode" (column 2) and add the collected data to list
    csv_datei13 = os.path.join(path, 'QuantitiyShare_electricity_from_Battery_to_CHPinStandby_Elect_Controller.csv')
    with open(csv_datei13, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 

            EnergyInputData13.append(row[1])
            iii +=1

    del iii


    # Load data from "Quantity share chp to battery" (column 2) and add the collected data to list
    csv_datei14 = os.path.join(path, 'QuantitiyShare_CHP_to_battery_Elect_Controller.csv')
    with open(csv_datei14, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1 
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData14.append(row[1])
            iii +=1
    del iii


    # Load data from "Quantity share pv to grid" (column 2) and add the collected data to list
    csv_datei15 = os.path.join(path, 'QuantitiyShare_CHP_to_grid_Elect_Controller.csv')
    with open(csv_datei15, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        iii = 1
        for row in csvreader:
            if iii > 1:
                row[1] = row[1].replace(',', '.')
                row[1] = float(row[1]) 
            EnergyInputData15.append(row[1])
            iii+=1
    del iii




    #Add all data in prepared list
    model_energy_input_data = [(x[0], x[1], x[2], EnergyInputData2[i], EnergyInputData3[i], EnergyInputData4[i], EnergyInputData5[i], EnergyInputData6[i], EnergyInputData7[i], EnergyInputData8[i], EnergyInputData9[i], EnergyInputData10[i], EnergyInputData11[i], EnergyInputData12[i], EnergyInputData13[i], EnergyInputData14[i], EnergyInputData15[i]) for i, x in enumerate(model_energy_input_data)]



    # Durchlaufen der Daten und Einfügen in die Excel-Datei
    for row_idx, row_data in enumerate(model_energy_input_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = None # First delete existing data in Cell
            worksheet.cell(row=row_idx, column=col_idx).value = value #integrate data in cell

 
    
    #workbook.save(path2)
    workbook.save(excelfilepathresults3)


def errormessage(daten):
    path = ResultPathProviderSingleton().get_result_directory_name() 
    name = "//ERRORMESSAGE.csv"
    path = path + name

    neue_daten = [daten]

    # Überprüfen, ob die CSV-Datei bereits existiert
    if os.path.exists(path):
        # CSV-Datei im Anhänge-Modus öffnen und neue Daten hinzufügen
        with open(path, mode='a', newline='') as file:
            csv_writer = csv.writer(file)
            
            # Informationen in die nächste freie Zeile der CSV-Datei anhängen
            csv_writer.writerow(neue_daten)
            
    else:
        # CSV-Datei existiert nicht, erstelle eine neue und füge Daten hinzu
        with open(path, mode='w', newline='') as file:
            csv_writer = csv.writer(file)
            
            # Informationen in die neue CSV-Datei schreiben
            csv_writer.writerow(neue_daten)
    

