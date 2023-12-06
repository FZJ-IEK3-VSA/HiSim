'''Cell4Life_Postprocessing.py
"C4L Specific" saving all Data'

'''
import csv
import os
import json
import openpyxl
import shutil  # Needed for copying of excel file
from datetime import datetime
from openpyxl.styles import NamedStyle
from hisim.result_path_provider import ResultPathProviderSingleton, SortingOptionEnum

def saveInputdata(input_variablen):
    PreResultNumber = input_variablen["PreResultNumber"]["value"]
    #----Save all Input Data in a .txt ----
    path = ResultPathProviderSingleton().get_result_directory_name() 
    name = f"//S{PreResultNumber}_InputConfig.csv"
    path = path + name
    del name

    with open(path, mode='w', newline='') as file:
        writer = csv.writer(file)
        # Schreibe den Header (Spaltenüberschriften)
        writer.writerow(['Variable', 'Wert', "Einheit"])
        # Schreibe die Variablen und ihre Werte in separate Zeilen
        for variable, info in input_variablen.items():
            writer.writerow([variable, info["value"], info["unit"]])
        file.close()

def saveexcelforevaluations(input_variablen, excelfilepathallresults, excel_filename):
    PreResultNumber = input_variablen["PreResultNumber"]["value"]
    path = ResultPathProviderSingleton().get_result_directory_name() 
    
    zusammengefuegte_daten = []
    Daten2 = []
    
    # Lade Daten aus dem ersten CSV-Datei "Electricity TO or FROM Grid Ergebnisse" (Spalte 1 und Spalte 2) und füge sie zur Liste hinzu
    erstes_csv_datei = os.path.join(path, 'ElectricityToOrFromGrid_L2EMSElectricityController.csv')
    with open(erstes_csv_datei, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            zusammengefuegte_daten.append([row[0], row[1]])

    # Lade Daten aus dem zweiten CSV-Datei "Thermal Output CHP in Watt" (ausschließlich Spalte 2) und füge sie zur Liste hinzu (in die dritte Spalte)
    zweites_csv_datei = os.path.join(path, 'ThermalPowerOutput_CHP_w2.csv')
    with open(zweites_csv_datei, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Daten2.append(row[1])
    

    #Zusammenfuehren der Daten
    zusammengefuegte_daten = [(x[0], x[1], Daten2[i]) for i, x in enumerate(zusammengefuegte_daten)]
    
    #------------
    # Create a complete new Excel File and save the result data in this resultpath/file
    workbook = openpyxl.Workbook()
    
    #Erstelle ein Excel-Arbeitsblatt
    worksheet = workbook.active
    worksheettitel = f"S{PreResultNumber}-Input std."
    worksheet.title = worksheettitel
    
    # Schreibe die zusammengeführten Daten in das Excel-Arbeitsblatt
    for row in zusammengefuegte_daten:
        worksheet.append(row)
    excelfilepath = path + '//' + worksheettitel + '.xlsx'
    workbook.save(excelfilepath)
    del excelfilepath, workbook, worksheet, row, worksheettitel

    #------------
    # Add all the result data from this simulation round within the economic assessment excel file; 
    workbook = openpyxl.load_workbook(excelfilepathallresults)
    # Wählen Sie das Tabellenblatt aus, in das Sie Ihre Daten kopieren möchten
    
    worksheettitel = 'S'+ str(PreResultNumber) + '-Input std.'
    worksheet = workbook[worksheettitel]
    del worksheettitel

    # Go through the following columns and delete the data
    zu_loeschende_spalten = [0, 1, 2]
    # Go through the rows and delete the data in the desired columns
    for zeile in range(worksheet.max_row):
        for spalte in zu_loeschende_spalten:
            zelle = worksheet.cell(row=zeile+1, column=spalte+1)
            zelle.value = None
    
    # Fügen Sie die neuen Daten an den gleichen Stellen ein, an denen die alten Daten gelöscht wurden
    for zeile, daten in enumerate(zusammengefuegte_daten, start=1):  # Start bei Zeile 2, um Überschriften zu vermeiden
        for spalte, wert in enumerate(daten, start=1):
            zelle = worksheet.cell(row=zeile, column=spalte)
            zelle.value = wert

    #------------    
    #Add Input Data to Economic Excel file
    # Select the desired worksheet
    worksheet = workbook['Anlagendaten']  # 'Anlagendaten' durch den tatsächlichen Namen ersetzen

    #In first simulation round, delete all data in worksheet "Anlagendaten"
    if PreResultNumber == 0:
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.value = None

    # Determine the row to append the data
    next_column = worksheet.max_column + 1
    next_row = 1
    # Write the data to the worksheet
    for parameter, info in input_variablen.items():
        if PreResultNumber == 0:
            worksheet.cell(row=next_row, column=1, value=parameter)
            worksheet.cell(row=next_row, column=2, value=info["unit"])
            worksheet.cell(row=next_row, column=3, value= info["value"])
        else:
            worksheet.cell(row=next_row, column=next_column, value= info["value"])    
        next_row += 1

    #------------
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("Parametervariation Runde: ", str(input_variablen["PreResultNumber"]["value"]))
    print("------------------------------------------------------------------------------")
    print("Parametervariation---Batteriekapazität: ", str(input_variablen["battery_capacity"]["value"]), "  &  Brennstoffzellenleistung: ",input_variablen["fuel_cell_power"]["value"])
    print("Simulation mit Parametern abgeschlossen - Ergebnisse werden im ökonomischen Bewertungstool gespeichert")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    # Save the Excel file
    workbook.save(excelfilepathallresults)



def makeacopyofevaluationfile(copytopath, filepath, name):
    #Make a copy of result file!
   
    #path = 'C://Users//Standard//C4LResults//results//'
    #filepath = path + 'OriginalExcelFile//20231019_oekonomische_Auswertung_v2.xlsx'
    
    # Erstellen Sie einen Datums- und Zeitstempel für die Sicherheitskopie
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    
    # Erstellen Sie den Dateinamen der Sicherheitskopie
    copyfile_filename = f'{name}_{timestamp}.xlsx'

    # Erstellen Sie den vollständigen Dateipfad für die Sicherheitskopie
    filepath_ofcopy = copytopath + copyfile_filename
    
    # Kopieren Sie die Excel-Datei, um die Sicherheitskopie zu erstellen
    shutil.copy2(filepath, filepath_ofcopy)
    
    return filepath_ofcopy, copyfile_filename




def save_data_in_excel_for_economic_assessment(input_variablen,excelfilepathresults):
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("Parametervariation Runde: ", str(input_variablen["PreResultNumber"]["value"]))
    print("------------------------------------------------------------------------------")
    print("Parametervariation---Batteriekapazität: ", str(input_variablen["battery_capacity"]["value"]), "  &  Brennstoffzellenleistung: ",input_variablen["fuel_cell_power"]["value"])
    print("Simulation mit Parametern abgeschlossen - Ergebnisse werden im ökonomischen Einzelfall-Bewertungstool gespeichert")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")
    print("------------------------------------------------------------------------------")

    PreResultNumber = input_variablen["PreResultNumber"]["value"]
    #----Save all Input Data in a .txt ----
    path = ResultPathProviderSingleton().get_result_directory_name() 
    workbook = openpyxl.load_workbook(excelfilepathresults)
    
    #----------------------------------------------   
    #Add "Anlagendaten" to excek
    worksheet = workbook['Anlagendaten']  # 'Anlagendaten' durch den tatsächlichen Namen ersetzen
    
    # delete all data in sheet

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.value = None
    
    next_column = worksheet.max_column + 1
    next_row = 1
    for parameter, info in input_variablen.items():
        worksheet.cell(row=next_row, column=1, value=parameter)
        worksheet.cell(row=next_row, column=2, value=info["unit"])
        worksheet.cell(row=next_row, column=3, value= info["value"])
        next_row += 1
    #----------------------------------------------
    
    #Add Input Data for calcuation within "Gesamtmodell"
    #Load Original Energy-Input Data
    model_energy_input_data = []
    EnergyInputData2 = []
    EnergyInputData3 = []
    EnergyInputData4 = []
    EnergyInputData5 = []

    # Load data from "-_Current" (column 1 & 2) and add the collected data to list
    csv_datei1 = os.path.join(path, '-_Current.csv')
    with open(csv_datei1, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        izaehler = 0
        for row in csvreader:
            if izaehler == 0:
                monat = 'Monat'
            else:
                datum = row[0]
                datum_obj = datetime.strptime(datum, '%Y-%m-%d %H:%M:%S')
                monat = datum_obj.month
            
            model_energy_input_data.append([row[0], monat, row[1]])
            izaehler += 1 


    # Load data from "CSV_PVComponent" (column 2) and add the collected data to list
    csv_datei2 = os.path.join(path, 'CSV_PVComponent.csv')
    with open(csv_datei2, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            EnergyInputData2.append(row[1])

    # Load data from "CSV_HeatingSystemComponent" (column 2) and add the collected data to list
    csv_datei3 = os.path.join(path, 'CSV_HeatingSystemComponent.csv')
    with open(csv_datei3, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            EnergyInputData3.append(row[1])

    # Load data from "CSV_WarmWaterComponent" (column 2) and add the collected data to list
    csv_datei4 = os.path.join(path, 'CSV_WarmWaterComponent.csv')
    with open(csv_datei4, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            EnergyInputData4.append(row[1])

    # Load data from "CSV_Sum_of_heat_energy_demand" (column 2) and add the collected data to list
    csv_datei5 = os.path.join(path, 'Sum_Sum_of_heat_energy_demand.csv')
    with open(csv_datei5, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            EnergyInputData5.append(row[1])


  


    #Add all data in prepared list
    model_energy_input_data = [(x[0], x[1], x[2], EnergyInputData2[i], EnergyInputData3[i], EnergyInputData4[i], EnergyInputData5[i]) for i, x in enumerate(model_energy_input_data)]
    #Save List in Excel File
    # Add all the result data from this simulation round within the economic assessment excel file; 
    worksheet = workbook['InputGesamtmodell']  # 'Anlagendaten' durch den tatsächlichen Namen ersetzen
    
    # Go through the following columns and delete the data
    zu_loeschende_spalten = [0, 1, 2, 3, 4, 5, 6,7]
    # Go through the rows and delete the data in the desired columns
    for zeile in range(worksheet.max_row):
        for spalte in zu_loeschende_spalten:
            zelle = worksheet.cell(row=zeile+1, column=spalte+1)
            zelle.value = None
    

    # Durchlaufen der Daten und Einfügen in die Excel-Datei
    for row_idx, row_data in enumerate(model_energy_input_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = value

    # # Excel-Berechnung für die 7. Spalte (Summe der ersten und zweiten Spalte)
    #     if row_idx == 1:
    #         worksheet.cell(row=row_idx, column=7).value = None
    #         worksheet.cell(row=row_idx, column=7).value = f"Electr. Result. = -Current needed + PV Component"  # Summe der Werte in der ersten und zweiten Spalte
    #     else: 
    #         worksheet.cell(row=row_idx, column=7).value = f"=-C{row_idx}+D{row_idx}"  # Summe der Werte in der ersten und zweiten Spalte
 

    del csv_datei1, csv_datei2, csv_datei3, csv_datei4, model_energy_input_data, EnergyInputData2, EnergyInputData3, EnergyInputData4, EnergyInputData5,izaehler
    #----------------------------------------------
    
    #Add Input Data for calcuation within "Gesamtmodell"
    #Prepare Model-Result-Data for Saving in Excel
    
    zusammengefuegte_daten = []
    Data2 = []
    Data3 = []
    Data4 = []
    Data5 = []
    Data6 = []
    Data7 = []
    Data8 = []
    Data9 = []
    Data10 = []
    Data11 = []


    # Lade Daten aus dem ersten CSV-Datei "Electricity TO or FROM Grid Ergebnisse" (Spalte 1 und Spalte 2) und füge sie zur Liste hinzu
    csv_datei1 = os.path.join(path, 'ElectricityToOrFromGrid_L2EMSElectricityController.csv')
    if not os.path.exists(csv_datei1):
        csv_datei1 = os.path.join(path, 'ElectricityToOrFromGrid_Elect_Controller.csv')


    with open(csv_datei1, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        # for row in csvreader:
        #     zusammengefuegte_daten.append([row[0], row[1]])

        izaehler = 0
        for row in csvreader:
            if izaehler == 0:
                monat = 'Monat'
            else:
                datum = row[0]
                datum_obj = datetime.strptime(datum, '%Y-%m-%d %H:%M:%S')
                monat = datum_obj.month
            
            zusammengefuegte_daten.append([row[0], monat, row[1]])
            izaehler += 1 

    # Lade Daten aus dem zweiten CSV-Datei "Thermal Output CHP in Watt" (ausschließlich Spalte 2) und füge sie zur Liste hinzu (in die dritte Spalte)
    csv_datei2 = os.path.join(path, 'ThermalPowerOutput_CHP_w2.csv')
    with open(csv_datei2, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data2.append(row[1])

    # Load 3. CSV-Datei "AcBatteryPower_Battery_w1" (only second column) und add it to the list
    csv_datei3 = os.path.join(path, 'AcBatteryPower_Battery_w1.csv')
    with open(csv_datei3, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data3.append(row[1])
            
    # Load 4. CSV-Datei "DcBatteryPower_Battery_w1" (only second column) und add it to the list
    csv_datei4 = os.path.join(path, 'DcBatteryPower_Battery_w1.csv')
    with open(csv_datei4, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data4.append(row[1])

    # Load 5. CSV-Datei "StateOfCharge_Battery_w1" (only second column) und add it to the list
    csv_datei5 = os.path.join(path, 'StateOfCharge_Battery_w1.csv')
    with open(csv_datei5, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data5.append(row[1])

    # Load 6. CSV-Datei "HydrogenSOC_HydrogenStorage_w999" (only second column) und add it to the list
    csv_datei6 = os.path.join(path, 'HydrogenSOC_HydrogenStorage_w999.csv')
    with open(csv_datei6, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data6.append(row[1])


#Neu:Electricity Consumption of different parts:
    # Load 7. data from "ElectricityConsumptionElectrolyzer" (column 2) and add the collected data to list
    csv_datei7 = os.path.join(path, 'ElectricityConsumptionElectrolyzer_StaticElectrolyzer.csv')
    with open(csv_datei7, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data7.append(row[1])

    # Load data from "StorageElectricityConsumption_HydrogenStorage_w999" (column 2) and add the collected data to list
    csv_datei8 = os.path.join(path, 'StorageElectricityConsumption_HydrogenStorage_w999.csv')
    with open(csv_datei8, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data8.append(row[1])

    # Load 9. CSV-Datei "TotalElectricityConsumption_L2EMSElectricityController.csv" (only second column) und add it to the list --> IS NOT THE TOTAL CONSUMPTION (based on a comparison and validation done by 4ward)
    csv_datei9 = os.path.join(path, 'TotalElectricityConsumption_Elect_Controller..csv')
    
    if not os.path.exists(csv_datei9):
        csv_datei9 = os.path.join(path, 'TotalElectricityConsumption_Elect_Controller.csv')

    with open(csv_datei9, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data9.append(row[1])

    # Load data from "from CHP to house" (column 2) and add the collected data to list
    csv_datei10 = os.path.join(path, 'QuantitiyShare_electricity_from_CHP_to_house_Elect_Controller.csv')
    with open(csv_datei10, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data10.append(row[1])

    # Load data from "from Battery to house" (column 2) and add the collected data to list
    csv_datei11 = os.path.join(path, 'QuantitiyShare_electricity_from_Battery_to_house_Elect_Controller.csv')
    with open(csv_datei11, 'r') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=';')  # Verwende Semikolon als Trennzeichen
        for row in csvreader:
            Data11.append(row[1])      

    #Zusammenfuehren der Daten
    zusammengefuegte_daten = [(x[0], x[1], x[2], Data2[i], Data3[i], Data4[i], Data5[i], Data6[i], Data7[i], Data8[i], Data9[i], Data10[i], Data11[i]) for i, x in enumerate(zusammengefuegte_daten)]

    

    #------------
    

    # Save "zusammengefuegte_data" which represents the Result data of the model in the excel file
    worksheet = workbook['OutputGesamtmodell']
    # delete all data in sheet
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.value = None

   
    for zeile, daten in enumerate(zusammengefuegte_daten, start=1):  # Start bei Zeile 2, um Überschriften zu vermeiden
        for spalte, wert in enumerate(daten, start=1):
            zelle = worksheet.cell(row=zeile, column=spalte)
            zelle.value = wert

    workbook.save(excelfilepathresults)


