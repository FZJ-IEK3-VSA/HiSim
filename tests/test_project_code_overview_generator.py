"""Test for project code overview generator."""
# clean
import logging
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from hisim import utils
from hisim.project_code_overview_generator import FileInformation, OverviewGenerator


@pytest.mark.extendedbase
@utils.measure_execution_time
def test_project_code_overview_generator(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Test that the project code overview generator produces its output.

    The :class:`OverviewGenerator` walks the HiSim source tree, collects
    information about every Python module and writes the result into an Excel
    workbook named ``components_information.xlsx``. Running it must therefore
    produce a non-empty workbook with the expected ``HiSim Files`` sheet.
    ``write_clean_files`` additionally emits the flake8/prospector call files
    into the parent of the working directory.

    ``run()`` writes its artifacts relative to the current working directory,
    so the test runs from a dedicated sub-folder of ``tmp_path``.  This keeps
    every artifact inside ``tmp_path`` (the ``../`` files land in ``tmp_path``
    itself) and avoids leaving calculation artifacts in the repository.
    """
    work_dir = tmp_path / "overview_work"
    work_dir.mkdir()
    monkeypatch.chdir(work_dir)

    overview_generator = OverviewGenerator()
    overview_generator.run()

    # Primary artifact: the Excel overview of all HiSim modules.
    expected_output = work_dir / "components_information.xlsx"
    assert expected_output.exists()
    assert expected_output.stat().st_size > 0

    workbook = load_workbook(filename=expected_output)
    assert "HiSim Files" in workbook.sheetnames
    worksheet = workbook["HiSim Files"]
    # The generator writes one block per source file; the repository contains
    # many modules, so the sheet must hold real tabular data.
    assert worksheet.max_row > 1
    assert worksheet.max_column > 1
    # The first column of the first row is the module name (a .py file name).
    first_module = worksheet.cell(row=1, column=1).value
    assert isinstance(first_module, str)
    assert first_module.endswith(".py")

    # Secondary artifacts written by write_clean_files into the parent dir.
    assert (tmp_path / "flake8_calls.txt").exists()
    assert (tmp_path / "prospector_calls.txt").exists()
    assert (tmp_path / "prospector_mass_call.cmd").exists()


@pytest.mark.base
def test_process_one_file_classifies_members_by_type(tmp_path: Path) -> None:
    """process_one_file classifies module members via inspect/isinstance checks.

    Regression test for replacing fragile ``str(type(x)) == "<class 'type'>"``
    string comparisons with ``inspect.isclass`` / ``inspect.ismodule`` /
    ``inspect.isfunction`` / ``isinstance``.  Classes that use a custom metaclass
    (a ``type`` subclass) -- e.g. pydantic models, Enums, or singleton classes --
    must be detected as classes, which the old string check silently missed.
    """
    sample = tmp_path / "test_overview_sample.py"
    sample.write_text(
        "# clean\n"
        '"""Test module for process_one_file type classification."""\n\n\n'
        "class MyMeta(type):\n"
        '    """Custom metaclass."""\n\n\n'
        "class RegularClass:\n"
        '    """A regular class."""\n\n\n'
        "class CustomMetaClass(metaclass=MyMeta):\n"
        '    """A class with a custom metaclass."""\n\n\n'
        "def my_function():\n"
        '    """A module-level function."""\n\n\n'
        '__authors__ = "Test Author"\n\n'
        'my_string = "hello"\n'
        "my_list = [1, 2, 3]\n"
        'my_dict = {"key": "value"}\n'
        "my_int = 42\n",
        encoding="utf-8",
    )

    generator = OverviewGenerator()
    # try_to_load_module registers the module in sys.modules under the file
    # basename (including the .py extension); clean up to avoid polluting
    # the module registry for subsequent tests.
    mod_key = sample.name
    try:
        myfi = generator.process_one_file(str(sample))
    finally:
        sys.modules.pop(mod_key, None)

    class_names = [c.class_name for c in myfi.classes]
    # RegularClass has the default 'type' metaclass -- detected by both old
    # and new code.
    assert "RegularClass" in class_names
    # CustomMetaClass uses a custom metaclass -- only detected by inspect.isclass.
    assert "CustomMetaClass" in class_names
    # MyMeta is itself a class (a metaclass subclass of type).
    assert "MyMeta" in class_names
    custom = next(c for c in myfi.classes if c.class_name == "CustomMetaClass")
    assert custom.lines_of_code > 0

    method_names = [m.method_name for m in myfi.methods]
    assert "my_function" in method_names

    # __authors__ is routed to the authors metadata field, not the strings list.
    assert myfi.authors == "Test Author"

    string_names = [s.string_name for s in myfi.strings]
    assert "my_string" in string_names

    list_names = [lst.list_name for lst in myfi.lists]
    assert "my_list" in list_names

    dict_names = [d.dict_name for d in myfi.dicts]
    assert "my_dict" in dict_names

    other_names = [o.member_name for o in myfi.others]
    assert "my_int" in other_names


@pytest.mark.base
def test_process_one_file_warns_on_duplicate_class_names(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """Duplicate class names across files log a warning instead of raising.

    With ``inspect.isclass`` the overview generator now correctly detects
    classes that use custom metaclasses (Enums, pydantic models, singletons).
    Several such classes share names across modules (e.g. ``IntEnum`` position
    enums).  Raising on the first duplicate would abort the entire run, so the
    check degrades to a logged warning while still recording every class.
    """
    file_a = tmp_path / "sample_a.py"
    file_b = tmp_path / "sample_b.py"
    content = "# clean\nclass SharedName:\n    pass\n"
    file_a.write_text(content, encoding="utf-8")
    file_b.write_text(content, encoding="utf-8")

    generator = OverviewGenerator()
    try:
        with caplog.at_level(logging.WARNING, logger="hisim.project_code_overview_generator"):
            myfi_a = generator.process_one_file(str(file_a))
            # Must not raise on the second occurrence of SharedName.
            myfi_b = generator.process_one_file(str(file_b))
    finally:
        sys.modules.pop(file_a.name, None)
        sys.modules.pop(file_b.name, None)

    assert "SharedName" in [c.class_name for c in myfi_a.classes]
    assert "SharedName" in [c.class_name for c in myfi_b.classes]
    assert any("SharedName" in record.getMessage() for record in caplog.records)


@pytest.mark.base
def test_try_to_load_module_returns_none_and_warns_when_spec_is_none(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """try_to_load_module fails loudly when no import spec can be created.

    For a path with no importable loader (e.g. a ``.txt`` file),
    :func:`importlib.util.spec_from_file_location` returns ``None``. Instead of
    letting ``module_from_spec`` raise a confusing ``AttributeError`` that the
    broad ``except`` would mask as a generic load failure, the method must
    return ``None`` and log a clear, specific warning so the unimportable file
    is visible. This guards the ``Optional[ModuleSpec]`` narrowing that removed
    the previous misleading ``# type: ignore`` suppressions.
    """
    not_a_module = tmp_path / "not_a_module.txt"
    not_a_module.write_text("not python", encoding="utf-8")

    myfi = FileInformation()
    myfi.file_name = str(not_a_module)
    myfi.module_name = not_a_module.name

    generator = OverviewGenerator()
    with caplog.at_level(logging.WARNING, logger="hisim.project_code_overview_generator"):
        result = generator.try_to_load_module(myfi)

    assert result is None
    assert any(
        "spec or loader is None" in record.getMessage()
        for record in caplog.records
    )


@pytest.mark.base
def test_process_one_file_records_class_without_source_definition(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """process_one_file records classes whose source cannot be located.

    ``inspect.getsourcelines`` raises ``OSError("could not find class
    definition")`` for classes created without a literal ``class`` statement
    (e.g. ``collections.namedtuple`` or the functional ``Enum`` API). A single
    such class must not abort the whole overview run; instead the class is
    recorded with a zero line count and a warning is logged so the
    unintrospectable class is visible rather than silently dropped.
    """
    module_file = tmp_path / "dynamic_classes.py"
    module_file.write_text(
        "# clean\n"
        "from collections import namedtuple\n"
        "# No 'class Point:' statement -> inspect cannot locate the source.\n"
        "Point = namedtuple('Point', ['x', 'y'])\n",
        encoding="utf-8",
    )

    generator = OverviewGenerator()
    try:
        with caplog.at_level(
            logging.WARNING, logger="hisim.project_code_overview_generator"
        ):
            myfi = generator.process_one_file(str(module_file))
    finally:
        sys.modules.pop(module_file.name, None)

    assert "Point" in [c.class_name for c in myfi.classes]
    point = next(c for c in myfi.classes if c.class_name == "Point")
    assert point.lines_of_code == 0
    assert any(
        "Point" in record.getMessage() for record in caplog.records
    )


@pytest.mark.base
def test_collect_files_only_walks_hisim_package() -> None:
    """collect_files must stay inside the ``hisim`` package directory.

    Regression test for the CI failure ``'array_api_backends' not found in
    `markers` configuration option``: ``collect_files`` walked
    ``__file__.parent.parent`` -- the repository root under an editable install,
    or ``site-packages`` under a non-editable install -- and imported every
    ``.py`` file it found, including third-party test modules (e.g. SciPy's
    ``test_array_api.py``) whose module-level ``pytest.mark.array_api_backends``
    raises ``pytest.Failed`` (a ``BaseException``) under ``--strict-markers``.
    The walk is scoped to the ``hisim`` package so only HiSim's own modules are
    ever imported.
    """
    import hisim.project_code_overview_generator as pcog

    generator = OverviewGenerator()
    files = generator.collect_files()
    assert files, "collect_files should find hisim package modules"
    hisim_pkg = Path(pcog.__file__).resolve().parent
    for f in files:
        resolved = Path(f).resolve()
        assert resolved.suffix == ".py"
        assert hisim_pkg in resolved.parents, (
            f"{f} is outside the hisim package directory {hisim_pkg}"
        )


@pytest.mark.base
def test_try_to_load_module_catches_base_exception_outcome(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """try_to_load_module skips modules that raise a ``BaseException`` outcome.

    Regression test for the CI failure ``'array_api_backends' not found in
    `markers` configuration option``: importing a module that accesses an
    unregistered ``pytest.mark`` under ``--strict-markers`` raises
    ``pytest.Failed``, which subclasses ``BaseException`` (not ``Exception``).
    The previous ``except (Exception, SystemExit)`` did not catch it, so the
    outcome escaped ``try_to_load_module`` and aborted the whole overview run.
    ``pytest.fail`` (which raises ``Failed``) stands in for that scenario: the
    unimportable file must be reported via a warning and skipped (return
    ``None``) instead of crashing the caller. ``KeyboardInterrupt`` is the one
    ``BaseException`` that must still propagate.
    """
    module_file = tmp_path / "raises_pytest_failed.py"
    module_file.write_text(
        "# clean\n"
        "import pytest\n"
        "pytest.fail('simulated unimportable module')\n",
        encoding="utf-8",
    )

    myfi = FileInformation()
    myfi.file_name = str(module_file)
    myfi.module_name = module_file.name

    generator = OverviewGenerator()
    try:
        with caplog.at_level(
            logging.WARNING, logger="hisim.project_code_overview_generator"
        ):
            result = generator.try_to_load_module(myfi)
    finally:
        sys.modules.pop(module_file.name, None)

    assert result is None
    assert any(
        "Could not load" in record.getMessage() and "Failed" in record.getMessage()
        for record in caplog.records
    )


@pytest.mark.base
def test_try_to_load_module_skips_module_with_unregistered_pytest_mark(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """try_to_load_module skips a module that accesses an unregistered marker.

    Regression test for the CI failure ``'array_api_backends' not found in
    `markers` configuration option``: a third-party test module (e.g. SciPy's
    ``test_array_api.py``) sets ``pytestmark = pytest.mark.array_api_backends``
    at module level. Under ``--strict-markers`` merely *accessing* that marker
    inside a running pytest session raises ``pytest.Failed`` (a
    ``BaseException``), so importing such a file via ``exec_module`` propagates
    the outcome. ``try_to_load_module`` must catch it, report the file visibly,
    and return ``None`` instead of aborting the overview run. This uses the
    exact trigger from the CI failure (not a ``pytest.fail`` stand-in) so a
    future change in how pytest signals unknown markers is caught here.
    """
    module_file = tmp_path / "test_array_api.py"
    module_file.write_text(
        "# clean\n"
        "import pytest\n"
        "pytestmark = pytest.mark.array_api_backends\n",
        encoding="utf-8",
    )

    myfi = FileInformation()
    myfi.file_name = str(module_file)
    myfi.module_name = module_file.name

    generator = OverviewGenerator()
    try:
        with caplog.at_level(
            logging.WARNING, logger="hisim.project_code_overview_generator"
        ):
            result = generator.try_to_load_module(myfi)
    finally:
        sys.modules.pop(module_file.name, None)

    assert result is None
    assert any(
        "Could not load" in record.getMessage()
        and "Failed" in record.getMessage()
        and "array_api_backends" in record.getMessage()
        for record in caplog.records
    )
