"""Converts the Excel worked-example workbooks into the YAML fixtures the tests read.

Implements cost-spec-v2 §3.3 (converter, validation rules 1-8, arithmetic cross-check),
§3.4 (YAML format), §3.6 (drift check) and §3.8 (review attestations).

The workbook is the source of truth; the YAML next to it is generated and must never be
hand-edited. Run this script after every workbook change::

    python tools/convert_worked_examples.py            # regenerate all YAML files
    python tools/convert_worked_examples.py --check    # CI: fail if a YAML drifted
    python tools/convert_worked_examples.py --list-unreviewed

Dependencies are deliberately limited to `openpyxl` and `PyYAML`: the drift check runs in a
bare CI job without HiSim installed, so this module must not import `hisim`.

**What this script protects.** The whole value of the worked-example library is that its
expected numbers are *tamper-evident*: they must not be able to drift quietly towards whatever
a buggy implementation happens to produce. This converter never computes an expected value. Each
one is the number Excel/LibreOffice cached when the author's own spreadsheet formula was last
recalculated, and that formula is a genuinely independent second implementation of the engine's
arithmetic (cost-spec-v2 §3). Four mechanisms keep it that way:

* the derivation travels with the number — every expected cell holds a formula, rewritten into
  named quantities, or an explicit note stating where the constant comes from (rule 5) — so a
  value pasted in from a failing test run stands out in the PR diff;
* pure-arithmetic formulas are re-evaluated in Python and compared against the cached value, which
  catches a workbook edited without recalculating it (§3.3 arithmetic cross-check);
* `--check` regenerates every YAML in CI and fails on a single byte of difference, so neither a
  hand-edited YAML nor an unconverted workbook edit survives review (§3.6);
* the whole semantic part of the YAML (inputs, derivations, expected values, tolerances) is hashed
  into a short content fingerprint, and the `review:` block is emitted only while the fingerprint
  the reviewer typed into the workbook still matches — so any semantic change invalidates the
  attestation by construction rather than by anyone remembering to clear a flag (§3.8).

Nothing here ever writes a workbook: openpyxl discards cached formula values on save, and attesting
a review has to stay a human action performed inside Excel with the sheet open (§3.8).

Workbook layout (see `tests/worked_examples/_template.xlsx`, §3.2). One sheet, three
sections, each opened by an all-caps marker in column A:

===============  ==========================  ==================  ==============
column A         column B                    column C            column D
===============  ==========================  ==================  ==============
``METADATA``     name / spec_section / computed_by / description / reviewed_by /
                 review_date / reviewed_fingerprint, one per row
``INPUTS``       literal value               comment             --
``EXPECTED``     formula (or noted constant) abs_tol             note
===============  ==========================  ==================  ==============

Every input and expected row carries a workbook-scoped defined name on its column-B cell, so
formulas read `=principal_in_euro*interest_rate` instead of `=B4*B5`.
"""

from __future__ import annotations

import argparse
import ast
import hashlib
import math
import operator
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple, Union

import yaml
from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Token, Tokenizer
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

#: Labels, in Excel and YAML alike, are valid Python identifiers in snake case (§3.2).
LABEL_PATTERN = re.compile(r"^[a-z][a-z0-9_]*$")

#: A content fingerprint as it is typed into the workbook and printed by this tool (§3.8).
FINGERPRINT_PATTERN = re.compile(r"^[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}$")

#: Section markers in column A. Uppercase, so they can never collide with a label.
METADATA_MARKER = "METADATA"
INPUTS_MARKER = "INPUTS"
EXPECTED_MARKER = "EXPECTED"
SECTION_MARKERS = (METADATA_MARKER, INPUTS_MARKER, EXPECTED_MARKER)

#: Metadata rows the template defines; the first four are mandatory (§3.2).
METADATA_KEYS = (
    "name",
    "spec_section",
    "computed_by",
    "description",
    "reviewed_by",
    "review_date",
    "reviewed_fingerprint",
)
REQUIRED_METADATA_KEYS = ("name", "spec_section", "computed_by", "description")
REVIEW_KEYS = ("reviewed_by", "review_date", "reviewed_fingerprint")

#: Rule 3: functions whose result changes between runs destroy cached-value reproducibility.
VOLATILE_FUNCTIONS = frozenset({"TODAY", "NOW", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT"})

#: Rule 8: worked examples use degenerate bands only, so band label triples are rejected (§3.2).
#: Only uncertainty vocabulary is caught here — "band" alone is legitimate tariff vocabulary
#: (a time-of-use band), so the check keys on the _min/_best_estimate/_max triple convention
#: instead. The legacy _avg spellings stay on the reject list so workbooks authored against the
#: pre-rename vocabulary are still caught.
BAND_LABEL_SUFFIXES = ("_min", "_best_estimate", "_max", "_minimum", "_maximum", "_avg", "_average")
BAND_LABEL_TOKENS = ("uncertainty", "_min_", "_best_estimate_", "_max_", "_avg_")

#: Sign convention (§3.2): these expected quantities are *timeline entries* of revenue type and
#: must therefore be negative or zero. Subsidy awards are deliberately absent: the solver reports
#: award amounts as positive magnitudes and only the evaluator mirrors them onto the timeline, so
#: a blanket "subsidy is negative" rule would be wrong for the `subsidies/` group.
REVENUE_LABEL_TOKENS = ("feed_in_revenue", "residual_value", "anyway_cost_credit", "loan_disbursement")

#: Relative tolerance of the arithmetic cross-check. LibreOffice writes full-precision floats, so
#: the two evaluations agree to round-off, not bit-for-bit.
ARITHMETIC_RELATIVE_TOLERANCE = 1e-9

#: Operators the pure-arithmetic re-evaluation understands (§3.3). Anything else (PMT, NPV, SUM,
#: MIN, ..., and Excel's postfix percent) is treated as a function-library cell whose cached
#: value is trusted as the independent reference.
ARITHMETIC_OPERATORS = frozenset({"+", "-", "*", "/", "^"})

#: Python equivalents of the Excel operators, for the cross-check evaluator.
ARITHMETIC_NODE_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Pow: operator.pow,
}

#: Everything before this line is hashed for the content fingerprint; the block itself is not.
REVIEW_BLOCK_MARKER = "\nreview:\n"

GENERATED_HEADER = (
    "# GENERATED from {source} by tools/convert_worked_examples.py.\n"
    "# Edit the workbook and re-run the converter; never edit this file by hand (§3.6).\n"
)


class ValidationError(Exception):
    """A workbook violates the authoring conventions (§3.2) or a validation rule (§3.3).

    The only error type this tool raises for author mistakes, so that `run()` can catch it per
    workbook, report every offending example in one pass and still exit non-zero. Its message
    always names the row (or the label) and the rule that was broken, because the reader is the
    author of the spreadsheet, not a Python developer.
    """


@dataclass
class ExpectedRow:
    """One row of the Expected table: a value, its tolerance and how it was derived.

    The unit of assertion of the whole library: `tests/test_worked_examples.py` runs the engine on
    the example's inputs and compares its result for `label` against `value` within `abs_tol`.
    Carrying `derivation` (the formula rewritten into named quantities) and `note` alongside the
    number is what makes a changed expectation reviewable — the diff shows both what changed and
    how it was computed (§3.4).
    """

    label: str
    value: float  # the cached spreadsheet value, in the quantity's own unit
    abs_tol: float  # absolute tolerance declared by the author in column C; never guessed
    derivation: Optional[str] = None  # Excel formula in named quantities; None for noted constants
    note: Optional[str] = None


@dataclass
class Example:
    """One converted workbook, ready to be emitted as YAML.

    The in-memory form of exactly one worked example — one workbook, one sheet, one YAML file
    (§3.1) — after parsing and validation but before rendering. `group` is the containing directory
    name (`financing`, `tariffs`, `subsidies`, `discounting`, `end_to_end`) and is not decoration:
    the test collector maps it to the engine entry point the example is run against (§3.5).
    """

    source_name: str  # workbook file name, echoed into the generated header
    group: str  # parent directory name = the calculator the example exercises
    metadata: Dict[str, str]  # every METADATA_KEYS entry, missing optional ones as ""
    inputs: List[Tuple[str, Union[float, int, str, bool]]] = field(default_factory=list)
    expected: List[ExpectedRow] = field(default_factory=list)


# --------------------------------------------------------------------------- helpers


def decimals_for_tolerance(tolerance: float) -> int:
    """Number of decimals a value is rounded to, derived from its declared tolerance (§3.3).

    An `abs_tol` of 0.01 means the example asserts cents, so printing more than two decimals in the
    YAML would suggest a precision the author never claimed — and would make the drift check
    sensitive to the last bits of a float. Capped at 12 decimals, which is well past double
    precision for the magnitudes worked examples use.

    Raises:
        ValidationError: If the tolerance is not positive (guards a blank or zeroed column C).
    """
    if tolerance <= 0:
        raise ValidationError("abs_tol must be > 0.")
    return min(12, max(0, int(math.ceil(-math.log10(tolerance)))))


def format_value(value: float, tolerance: float) -> str:
    """Formats an expected value with the decimals its tolerance implies.

    Together with the other `format_*` helpers this is what makes the generated YAML a *stable*
    text: the same workbook must always render byte-for-byte identically, otherwise the CI drift
    check (§3.6) and the content fingerprint (§3.8) would fire on rounding noise instead of on real
    edits. The negative-zero guard exists for the same reason — a value that rounds to zero must
    render the same regardless of which side it approached from.
    """
    decimals = decimals_for_tolerance(tolerance)
    text = f"{round(value, decimals):.{decimals}f}"
    if text.startswith("-") and float(text) == 0.0:
        text = text[1:]  # never emit "-0.00"
    return text


def format_tolerance(tolerance: float) -> str:
    """Formats a tolerance in plain decimal notation (YAML 1.1 needs no exponent surprises).

    A tolerance written as `1e-2` is legal YAML 1.2 but is parsed as a *string* by YAML 1.1
    loaders, which would silently turn a numeric comparison in the collector into a type error.
    At least one decimal is always emitted so the value also reads unambiguously as a float.
    """
    decimals = max(1, decimals_for_tolerance(tolerance))
    return f"{tolerance:.{decimals}f}"


def format_number(value: Union[float, int]) -> str:
    """Formats an input number deterministically and without exponent notation.

    Inputs are transcribed rather than rounded — they are what the author typed and what the test
    feeds back into the engine — so this keeps full precision while removing the two ways Python
    can render the same value differently: exponent notation and a trailing `.0` on integral
    floats. Deterministic rendering is a requirement of the drift check, not a cosmetic choice.
    """
    if isinstance(value, int):
        return str(value)
    if float(value).is_integer() and abs(value) < 1e15:
        return str(int(value))
    text = repr(float(value))
    if "e" in text or "E" in text:
        text = f"{value:.15f}".rstrip("0")
        if text.endswith("."):
            text += "0"
    return text


def quote(text: str) -> str:
    """Double-quoted YAML scalar; newlines from Excel cells collapse to spaces.

    Every string the workbook contributes (names, descriptions, derivations, notes) goes through
    here, so that a cell containing a colon, a leading `#` or an alt-enter line break cannot break
    the emitted document or turn into a multi-line block whose exact indentation would then be part
    of the fingerprint. Whitespace is collapsed for the same reason: it must not carry meaning.
    """
    collapsed = " ".join(str(text).split())
    escaped = collapsed.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def format_scalar(value: Union[float, int, str, bool]) -> str:
    """Formats an input value (number, string or boolean) as a YAML scalar.

    The single entry point for rendering the INPUTS table, so that the type openpyxl read out of
    the cell survives into the fixture: numbers stay numbers, strings are quoted, and booleans
    render as YAML booleans. The bool branch comes first because `bool` is a subclass of `int` in
    Python and would otherwise be printed as 0/1.
    """
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return format_number(value)
    return quote(str(value))


def content_for_fingerprint(yaml_text: str) -> str:
    """The YAML minus its `review:` block — the content the fingerprint addresses (§3.8).

    Excluding the attestation from what it attests to is what breaks the self-reference: the block
    the reviewer's signature lands in would otherwise change the hash it contains. Everything
    semantic (inputs, derivations, expected values, tolerances) lies before the marker and is
    therefore inside the hash, so any edit to the example invalidates the attestation automatically.

    `tests/test_worked_examples.py` re-implements this same split from the YAML alone (it never
    opens a workbook), so the marker string must stay identical on both sides.
    """
    index = yaml_text.find(REVIEW_BLOCK_MARKER)
    if index == -1:
        return yaml_text
    return yaml_text[: index + 1]


def fingerprint_of(content: str) -> str:
    """Shortened SHA-256 of the fingerprinted content, displayed as XXXX-XXXX-XXXX (§3.8).

    The attestation token: the converter prints it, the reviewer types it into the workbook's
    `reviewed_fingerprint` row after re-deriving the example by hand, and from then on it is
    reproduced by anyone regenerating the YAML. Truncated to 12 hex characters and grouped in
    threes purely for human transcription — this is a change detector, not a security boundary.
    """
    digest = hashlib.sha256(content.encode("utf-8")).hexdigest()[:12].upper()
    return f"{digest[0:4]}-{digest[4:8]}-{digest[8:12]}"


# --------------------------------------------------------------------------- workbook parsing


def _cell_text(value: object) -> str:
    """Trimmed string content of a cell; empty string for blanks.

    Normalizes the two ways a spreadsheet says "nothing here" — an absent cell (`None`) and a cell
    holding only spaces — into one, because section detection and metadata parsing both treat blank
    and missing identically. Non-string cell values are stringified, so a numeric label or a
    date-formatted `review_date` still reaches the caller as text.
    """
    if value is None:
        return ""
    return str(value).strip()


def _defined_name_targets(workbook) -> Dict[str, str]:
    """Maps every workbook-scoped defined name to the coordinate it points at.

    This map is the backbone of the whole converter: it is what lets validation rule 1 ("every
    referenced cell has a name") be enforced and what turns click-and-point formulas such as
    `=B4*B5` into the readable `=principal_in_euro*interest_rate` that ships as the derivation.
    The sheet part and the `$` anchors are stripped because a name's target is compared against
    plain coordinates like `B7` everywhere downstream.
    """
    targets: Dict[str, str] = {}
    for name, definition in workbook.defined_names.items():
        attr_text = str(definition.attr_text or "")
        _, _, reference = attr_text.rpartition("!")
        targets[name] = reference.replace("$", "").upper()
    return targets


def _collect_sections(sheet) -> Dict[str, List[Tuple[int, List[object]]]]:
    """Splits the sheet into its METADATA / INPUTS / EXPECTED row blocks.

    Implements the template layout of §3.2 with the minimum of structure an author can get wrong:
    an all-caps marker in column A opens a section and the next blank column-A cell closes it, so
    the sections may sit anywhere on the sheet and be spaced out freely. Row numbers are carried
    along with the cells because every later error message, and the defined-name check, refer to
    the spreadsheet row the author is looking at.

    Returns:
        One list of `(row_index, [A, B, C, D] values)` per marker, in sheet order.

    Raises:
        ValidationError: If a labeled row appears before any section marker — usually a stray note
            typed into column A, which would otherwise be silently dropped from the fixture.
    """
    sections: Dict[str, List[Tuple[int, List[object]]]] = {marker: [] for marker in SECTION_MARKERS}
    current: Optional[str] = None
    for row_index in range(1, sheet.max_row + 1):
        label = _cell_text(sheet.cell(row_index, 1).value)
        if label in SECTION_MARKERS:
            current = label
            continue
        if not label:
            current = None  # a blank row in column A closes the current section
            continue
        if current is None:
            raise ValidationError(f"row {row_index}: label {label!r} outside any section marker.")
        cells = [sheet.cell(row_index, column).value for column in range(1, 5)]
        sections[current].append((row_index, cells))
    return sections


def _check_label(label: str, row_index: int, seen: Dict[str, int]) -> None:
    """Validation rules 2, 6 and 8 for one label.

    Labels are the one identifier that has to line up across three layers — Excel defined name,
    YAML key, Python field name — so a reviewer can trace a quantity through all of them by name
    (§3.2). Hence the identifier pattern (rule 6) and the workbook-wide uniqueness check (rule 2,
    shared between the INPUTS and EXPECTED tables via the mutable `seen` map, which this function
    updates in place). Rule 8 additionally rejects anything that looks like uncertainty-band
    syntax, because worked examples are authored with degenerate bands only.

    Args:
        label: The column-A text of the row.
        row_index: Spreadsheet row, used in the error messages and recorded in `seen`.
        seen: Labels already accepted in this workbook, mapped to their row; extended here.

    Raises:
        ValidationError: If the label is not an identifier, duplicates an earlier row, or carries
            min/best_estimate/max band vocabulary.
    """
    if not LABEL_PATTERN.match(label):
        raise ValidationError(f"row {row_index}: label {label!r} is not a valid identifier (rule 6).")
    if label in seen:
        raise ValidationError(f"row {row_index}: label {label!r} duplicates row {seen[label]} (rule 2).")
    if label.endswith(BAND_LABEL_SUFFIXES) or any(token in label for token in BAND_LABEL_TOKENS):
        raise ValidationError(
            f"row {row_index}: label {label!r} looks like uncertainty-band syntax; worked examples "
            "use degenerate bands only (rule 8)."
        )
    seen[label] = row_index


# --------------------------------------------------------------------------- formula handling


def _split_reference(reference: str) -> Tuple[Optional[str], str]:
    """Splits `'Sheet'!$B$5` into (sheet, B5); returns (None, ref) for a plain reference.

    Callers need the two halves separately: the sheet part is what rule 4 checks (one example is
    one sheet, so a cross-sheet reference is an error), while the normalized coordinate is what the
    name lookup keys on. Absolute-reference `$` markers and letter case are stripped so that `$B$5`,
    `B$5` and `b5` all resolve to the same cell.
    """
    if "!" not in reference:
        return None, reference.replace("$", "").upper()
    sheet, _, cell = reference.rpartition("!")
    return sheet.strip("'"), cell.replace("$", "").upper()


def _range_coordinates(start: str, end: str) -> List[str]:
    """Every coordinate of a rectangular range, row-major.

    Used only to enforce rule 7: a `SUM(B5:B14)` may be rewritten to
    `SUM(cash_flow_year_1:cash_flow_year_10)` only if *every* cell in between carries a name, so
    the expansion has to be materialized to check them one by one. Endpoints are normalized so the
    range may be written in either direction.
    """
    start_row, start_column = coordinate_to_tuple(start)
    end_row, end_column = coordinate_to_tuple(end)
    coordinates = []
    for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
        for column in range(min(start_column, end_column), max(start_column, end_column) + 1):
            coordinates.append(f"{get_column_letter(column)}{row}")
    return coordinates


def _rewrite_operand(
    reference: str,
    name_by_coordinate: Dict[str, str],
    known_names: Sequence[str],
    sheet_title: str,
    label: str,
) -> str:
    """Rewrites one operand token to defined names, applying rules 1, 4 and 7.

    This is where "no magic cells" is enforced: an operand that resolves to a cell without a
    defined name aborts the conversion, so every quantity in a derivation chain is named, visible
    in the Expected/Inputs tables and traceable by a reviewer. Operands already written as a name
    pass through untouched, which is the normal case for authors following §3.2.

    Args:
        reference: The raw operand text of the token (`B7`, `$B$7`, `'Sheet1'!B7`, `B5:B14`, or an
            already-defined name).
        name_by_coordinate: Coordinate to defined name, from `_defined_name_targets`.
        known_names: All defined names of the workbook, used to detect the pass-through case.
        sheet_title: The example's only sheet; anything else is a rule-4 violation.
        label: Label of the row being converted, for the error messages.

    Returns:
        The operand rewritten in terms of defined names.

    Raises:
        ValidationError: On an external-workbook reference or a foreign sheet (rule 4), a range
            covering an unnamed cell (rule 7), or an unnamed single cell (rule 1).
    """
    if "[" in reference:
        raise ValidationError(f"{label}: external workbook reference {reference!r} (rule 4).")
    if reference in known_names:
        return reference  # already written in terms of a defined name
    sheet, cell = _split_reference(reference)
    if sheet is not None and sheet != sheet_title:
        raise ValidationError(f"{label}: reference to another sheet ({reference!r}); one sheet per example (rule 4).")
    if ":" in cell:
        start, _, end = cell.partition(":")
        for coordinate in _range_coordinates(start, end):
            if coordinate not in name_by_coordinate:
                raise ValidationError(
                    f"{label}: range {reference!r} covers unnamed cell {coordinate}; ranges must consist "
                    "of individually labeled rows (rule 7)."
                )
        return f"{name_by_coordinate[start]}:{name_by_coordinate[end]}"
    if cell not in name_by_coordinate:
        raise ValidationError(
            f"{label}: formula references cell {reference!r}, which carries no name (rule 1)."
        )
    return name_by_coordinate[cell]


def rewrite_formula(
    formula: str,
    name_by_coordinate: Dict[str, str],
    known_names: Sequence[str],
    sheet_title: str,
    label: str,
) -> Tuple[str, List[Token]]:
    """Rewrites raw references to names and returns (derivation text, token stream).

    Excel surface syntax is preserved on purpose (§3.3): `^`, `PMT`, `SUM` and friends stay as
    they are, because the spreadsheet formula *is* the human-readable derivation.

    The rewrite runs over `openpyxl`'s formula tokenizer rather than a regular expression, so that
    a cell coordinate inside a string literal or a function name can never be mistaken for a
    reference. The returned token stream is what the caller inspects to decide whether the formula
    is pure arithmetic and can therefore be re-evaluated as a stale-cache check.

    Args:
        formula: The raw cell formula including its leading `=`.
        name_by_coordinate: Coordinate to defined name, from `_defined_name_targets`.
        known_names: All defined names of the workbook.
        sheet_title: Title of the example's only sheet.
        label: Label of the row being converted, for the error messages.

    Returns:
        `(derivation, tokens)` — the formula in named quantities as it will appear in the YAML
        (without the leading `=`, since the tokenizer drops it), and the raw token list.

    Raises:
        ValidationError: On a volatile function (rule 3), or via `_rewrite_operand` on rules 1, 4
            and 7.
    """
    tokens = Tokenizer(formula).items
    pieces: List[str] = []
    for token in tokens:
        if token.type == Token.FUNC and token.subtype == Token.OPEN:
            function_name = token.value.rstrip("(").upper()
            if function_name in VOLATILE_FUNCTIONS:
                raise ValidationError(f"{label}: volatile function {function_name}() is not allowed (rule 3).")
            pieces.append(token.value)
        elif token.type == Token.FUNC and token.subtype == Token.CLOSE:
            pieces.append(")")
        elif token.type == Token.PAREN:
            pieces.append("(" if token.subtype == Token.OPEN else ")")
        elif token.type == Token.SEP:
            pieces.append("," if token.subtype == Token.ARG else ";")
        elif token.type == Token.OPERAND and token.subtype == Token.TEXT:
            pieces.append('"' + str(token.value).replace('"', '""') + '"')
        elif token.type == Token.OPERAND and token.subtype == Token.RANGE:
            pieces.append(_rewrite_operand(str(token.value), name_by_coordinate, known_names, sheet_title, label))
        else:
            pieces.append(str(token.value))
    return "".join(pieces), list(tokens)


def _is_pure_arithmetic(tokens: Sequence[Token]) -> bool:
    """Whether the formula uses only `+ - * / ^ ( )` over names and literal numbers (§3.3).

    Decides which expected cells get the arithmetic cross-check. Cells that call into Excel's
    function library (`PMT`, `NPV`, `SUM`, ...) deliberately do not: those are exactly the cells
    whose Excel implementation is wanted as the independent reference, so re-deriving them in
    Python would defeat the purpose of the library. Excel's postfix `%` also falls out here,
    since it is not in `ARITHMETIC_OPERATORS`.
    """
    for token in tokens:
        if token.type == Token.FUNC:
            return False
        if token.type in (Token.ARRAY, Token.SEP):
            return False
        if token.type == Token.OPERAND and token.subtype not in (Token.RANGE, Token.NUMBER):
            return False
        if token.type in (Token.OP_PRE, Token.OP_IN, Token.OP_POST) and str(token.value) not in ARITHMETIC_OPERATORS:
            return False
    return True


def _arithmetic_expression(tokens: Sequence[Token], values: Dict[str, float], label: str) -> str:
    """Translates a pure-arithmetic token stream into an equivalent Python expression.

    Each named operand is substituted by its cached numeric value (via `repr`, so no precision is
    lost in the round trip) and Excel's `^` becomes Python's `**`; everything else is passed
    through verbatim. The result is a string of numbers and operators only — it contains no names
    at all — which is what allows `_evaluate_arithmetic` to parse it without any notion of scope.

    **Unary minus is parenthesized, because the two languages disagree about it.** Excel binds a
    leading `-` *tighter* than `^`, so `=-A1^2` means `(-A1)^2`; Python binds `**` tighter than
    unary minus, so the transliterated `-A1**2` would mean `-(A1**2)` — same text, opposite sign
    for an odd power, and a cross-check failure with a message about a stale workbook that is not
    stale. Every prefix `+`/`-` therefore gets a parenthesis that closes as soon as its operand is
    complete (a number, a substituted value, or a balanced parenthesized group), which reproduces
    Excel's binding exactly and is a no-op wherever the two languages already agreed.

    Args:
        tokens: Token stream of the already-rewritten formula (names, not coordinates).
        values: Cached value per defined name, collected up front in `convert_workbook`.
        label: Label of the row, for the error message.

    Returns:
        A Python expression string equivalent to the Excel formula.

    Raises:
        ValidationError: If an operand has no cached numeric value — e.g. a formula referencing a
            text input, which cannot be cross-checked.
    """
    pieces: List[str] = []
    pending_unary: List[int] = []
    depth = 0

    def close_completed_unaries() -> None:
        """Closes every prefix operator whose operand has just been fully emitted."""
        while pending_unary and pending_unary[-1] == depth:
            pieces.append(")")
            pending_unary.pop()

    for token in tokens:
        value = str(token.value)
        if token.type == Token.OPERAND and token.subtype == Token.RANGE:
            if value not in values:
                raise ValidationError(f"{label}: cannot re-evaluate {value!r}; it has no numeric value.")
            pieces.append(f"({values[value]!r})")
            close_completed_unaries()
        elif token.type == Token.OPERAND:
            pieces.append(value)
            close_completed_unaries()
        elif token.type == Token.PAREN:
            if token.subtype == Token.OPEN:
                pieces.append("(")
                depth += 1
            else:
                pieces.append(")")
                depth -= 1
                close_completed_unaries()
        elif token.type == Token.OP_PRE:
            pieces.append("(" + value)
            pending_unary.append(depth)
        elif token.type in (Token.OP_IN, Token.OP_POST):
            pieces.append("**" if value == "^" else value)
        else:
            pieces.append(" " if token.type == Token.WSPACE else value)
    return "".join(pieces)


def _evaluate_node(node: ast.AST) -> float:
    """Evaluates one node of the arithmetic expression tree; rejects everything else.

    A deliberately tiny interpreter over the four literal/operator node types the cross-check can
    encounter: numeric constants, unary +/-, and the five binary operators of
    `ARITHMETIC_NODE_OPERATORS`. Anything else — a name, a call, a comparison, a subscript — raises
    instead of being evaluated, which is what makes reading a workbook safe even though the
    expression text is derived from file content.
    """
    if isinstance(node, ast.Expression):
        return _evaluate_node(node.body)
    if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)) and not isinstance(node.value, bool):
        return float(node.value)
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
        value = _evaluate_node(node.operand)
        return value if isinstance(node.op, ast.UAdd) else -value
    if isinstance(node, ast.BinOp) and type(node.op) in ARITHMETIC_NODE_OPERATORS:
        return float(ARITHMETIC_NODE_OPERATORS[type(node.op)](_evaluate_node(node.left), _evaluate_node(node.right)))
    raise ValidationError(f"unsupported element {type(node).__name__} in the arithmetic cross-check.")


def _evaluate_arithmetic(expression: str) -> float:
    """Evaluates a numbers-and-operators expression without ever calling `eval` (§3.3).

    Parsing to an AST and walking it with `_evaluate_node` gives the same arithmetic as `eval`
    while making it structurally impossible for spreadsheet content to execute code. The check
    relies on Python's precedence matching Excel's for the supported operators, with one
    difference that `_arithmetic_expression` neutralizes by parenthesizing before the text ever
    reaches here: Excel binds unary minus tighter than `^` and Python does not. One difference
    remains unhandled — Excel's `^` is left-associative (`2^3^2` is 64) while Python's `**` is
    right-associative (512) — and it is left that way deliberately: a chained exponentiation
    surfaces as a cross-check *failure*, never as a silently accepted value, which is the
    conservative direction for a check whose whole purpose is catching wrong numbers.
    """
    return _evaluate_node(ast.parse(expression, mode="eval"))


def _cross_check_arithmetic(
    derivation_tokens: Sequence[Token],
    named_values: Dict[str, float],
    cached: float,
    label: str,
) -> None:
    """Re-evaluates a pure-arithmetic formula in Python; stale-cache defense of §3.3.

    The converter reads a workbook's *cached* values, which are only as fresh as the last
    recalculation — so an author who edits an input and saves without letting the spreadsheet
    recompute would otherwise export stale expectations that look perfectly authoritative. Where
    the formula is simple enough to redo independently, this catches exactly that, and it does so
    at conversion time rather than in a test failure nobody can interpret.

    Args:
        derivation_tokens: Tokens of the rewritten formula (`Tokenizer("=" + derivation)`).
        named_values: Cached value per defined name of the workbook.
        cached: The cached value of the cell under check.
        label: Label of the row, for the error message.

    Raises:
        ValidationError: If the formula cannot be re-evaluated, or if the two evaluations disagree
            by more than `ARITHMETIC_RELATIVE_TOLERANCE` relative to the larger magnitude.
    """
    expression = _arithmetic_expression(derivation_tokens, named_values, label)
    try:
        recomputed = _evaluate_arithmetic(expression)
    except ZeroDivisionError as error:
        raise ValidationError(f"{label}: re-evaluating the formula divides by zero.") from error
    except (SyntaxError, TypeError, ValueError) as error:
        raise ValidationError(f"{label}: cannot re-evaluate the formula ({error}).") from error
    scale = max(1.0, abs(recomputed), abs(cached))
    if abs(recomputed - cached) > ARITHMETIC_RELATIVE_TOLERANCE * scale:
        raise ValidationError(
            f"{label}: the cached Excel value {cached!r} disagrees with a Python re-evaluation of the "
            f"same formula ({recomputed!r}). The workbook was most likely edited without recalculating "
            "it (§3.3 arithmetic cross-check)."
        )


# --------------------------------------------------------------------------- conversion


def _parse_metadata(rows: Sequence[Tuple[int, List[object]]]) -> Dict[str, str]:
    """Reads the metadata block and validates §3.2 completeness plus the §3.8 review triple.

    Metadata is what makes an example reviewable rather than merely runnable: `spec_section` says
    which rule it pins down, `computed_by` says who (or which tool) produced the numbers, and
    `description` states the derivation in words. All four mandatory fields therefore have to be
    present before anything is emitted. The three review rows are all-or-nothing on purpose — a
    half-filled attestation is ambiguous, and an unparsable fingerprint would silently downgrade a
    reviewed example to "unreviewed" instead of being reported.

    Args:
        rows: The METADATA section rows from `_collect_sections`.

    Returns:
        Every key of `METADATA_KEYS`, with unfilled optional rows as empty strings.

    Raises:
        ValidationError: On an unknown or repeated key, a missing mandatory field, a partially
            filled review triple, or a fingerprint not of the form XXXX-XXXX-XXXX.
    """
    metadata: Dict[str, str] = {key: "" for key in METADATA_KEYS}
    seen: Dict[str, int] = {}
    for row_index, cells in rows:
        key = _cell_text(cells[0])
        if key not in METADATA_KEYS:
            raise ValidationError(f"row {row_index}: unknown metadata key {key!r} (expected one of {METADATA_KEYS}).")
        if key in seen:
            raise ValidationError(f"row {row_index}: metadata key {key!r} appears twice (rule 2).")
        seen[key] = row_index
        metadata[key] = _cell_text(cells[1])
    for key in REQUIRED_METADATA_KEYS:
        if not metadata[key]:
            raise ValidationError(f"metadata field {key!r} is mandatory and must not be empty (§3.2).")
    filled = [key for key in REVIEW_KEYS if metadata[key]]
    if filled and len(filled) != len(REVIEW_KEYS):
        raise ValidationError(
            f"review attestation is incomplete: {filled} set, {[k for k in REVIEW_KEYS if not metadata[k]]} empty. "
            "All three rows are set together or all three stay empty (§3.8)."
        )
    if metadata["reviewed_fingerprint"] and not FINGERPRINT_PATTERN.match(metadata["reviewed_fingerprint"]):
        raise ValidationError(
            f"reviewed_fingerprint {metadata['reviewed_fingerprint']!r} is not of the form XXXX-XXXX-XXXX (§3.8)."
        )
    return metadata


def convert_workbook(path: str) -> Tuple[Example, str]:
    """Converts one workbook into an :class:`Example` and its generated YAML text.

    The heart of the tool, and the place where all eight validation rules of §3.3 are actually
    applied to a sheet. The workbook is opened twice — once for formula text and once for the
    cached values Excel/LibreOffice last computed — because the fixture needs both: the cached
    number is the expectation, the formula is the derivation that makes it auditable. Nothing is
    written to disk here; `run()` decides whether the returned text is saved or only compared.

    The INPUTS table must hold literal values (a formula there would mean the example's own input
    is derived from something invisible), while every EXPECTED row must hold a formula or an
    explicit note, a positive tolerance, and — for revenue-type quantities — the correct sign
    under the engine's cost-positive/revenue-negative convention.

    Args:
        path: Path to the `.xlsx` file; its parent directory name becomes the example's group.

    Returns:
        `(example, yaml_text)` — the parsed example and the exact text that belongs in the `.yaml`
        next to the workbook.

    Raises:
        ValidationError: On any §3.1/§3.2 layout violation (multiple sheets, sheet-scoped names,
            doubly named cells), any of the eight validation rules, a failed arithmetic
            cross-check, a missing cached value, or an empty INPUTS/EXPECTED table.
    """
    source_name = os.path.basename(path)
    group = os.path.basename(os.path.dirname(path))
    formulas = load_workbook(path, data_only=False)
    values = load_workbook(path, data_only=True)
    if len(formulas.worksheets) != 1:
        raise ValidationError(
            f"{source_name}: workbook holds {len(formulas.worksheets)} sheets; one example is one "
            "workbook with one sheet (§3.1)."
        )
    formula_sheet = formulas.worksheets[0]
    value_sheet = values[formula_sheet.title]

    name_targets = _defined_name_targets(formulas)
    if formula_sheet.defined_names:
        raise ValidationError(
            f"{source_name}: sheet-scoped defined names {sorted(formula_sheet.defined_names)} found; names "
            "must be workbook-scoped (§3.1)."
        )
    name_by_coordinate: Dict[str, str] = {}
    for name, coordinate in name_targets.items():
        if coordinate in name_by_coordinate:
            raise ValidationError(
                f"{source_name}: cell {coordinate} carries two names "
                f"({name_by_coordinate[coordinate]!r} and {name!r}) (rule 2)."
            )
        name_by_coordinate[coordinate] = name

    sections = _collect_sections(formula_sheet)
    metadata = _parse_metadata(sections[METADATA_MARKER])
    example = Example(source_name=source_name, group=group, metadata=metadata)
    seen_labels: Dict[str, int] = {}
    # Cached values of every named cell, resolved up front: formulas may reference rows that
    # appear further down the sheet (year N interest referring to year N-1 remaining debt).
    named_values: Dict[str, float] = {}
    for coordinate, name in name_by_coordinate.items():
        if ":" in coordinate:
            continue
        cell_row, cell_column = coordinate_to_tuple(coordinate)
        cell_value = value_sheet.cell(cell_row, cell_column).value
        if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
            named_values[name] = float(cell_value)

    for row_index, cells in sections[INPUTS_MARKER]:
        label = _cell_text(cells[0])
        _check_label(label, row_index, seen_labels)
        coordinate = f"B{row_index}"
        if name_by_coordinate.get(coordinate) != label:
            raise ValidationError(
                f"row {row_index}: input {label!r} has no defined name on {coordinate} "
                f"(found {name_by_coordinate.get(coordinate)!r}); every quantity must be named (rule 1)."
            )
        raw = formula_sheet.cell(row_index, 2).value
        if isinstance(raw, str) and raw.startswith("="):
            raise ValidationError(f"row {row_index}: input {label!r} holds a formula; inputs are literal values.")
        if raw is None:
            raise ValidationError(f"row {row_index}: input {label!r} has no value.")
        if isinstance(raw, str) and re.match(r"^\s*-?[\d.]+\s*/\s*-?[\d.]+\s*/\s*-?[\d.]+\s*$", raw):
            raise ValidationError(f"row {row_index}: input {label!r} looks like a min/best_estimate/max triple (rule 8).")
        example.inputs.append((label, raw))

    for row_index, cells in sections[EXPECTED_MARKER]:
        label = _cell_text(cells[0])
        _check_label(label, row_index, seen_labels)
        coordinate = f"B{row_index}"
        if name_by_coordinate.get(coordinate) != label:
            raise ValidationError(
                f"row {row_index}: expected value {label!r} has no defined name on {coordinate} "
                f"(found {name_by_coordinate.get(coordinate)!r}) (rule 1)."
            )
        cached = value_sheet.cell(row_index, 2).value
        if not isinstance(cached, (int, float)) or isinstance(cached, bool):
            raise ValidationError(
                f"row {row_index}: expected value {label!r} has no cached number "
                f"(found {cached!r}); recalculate the workbook in Excel/LibreOffice before converting."
            )
        tolerance_raw = cells[2]
        if not isinstance(tolerance_raw, (int, float)) or isinstance(tolerance_raw, bool) or tolerance_raw <= 0:
            raise ValidationError(
                f"row {row_index}: expected value {label!r} needs a positive numeric tolerance in column C "
                f"(found {tolerance_raw!r}) (rule 6)."
            )
        note = _cell_text(cells[3]) or None
        raw_formula = formula_sheet.cell(row_index, 2).value
        derivation: Optional[str] = None
        if isinstance(raw_formula, str) and raw_formula.startswith("="):
            derivation, tokens = rewrite_formula(
                raw_formula, name_by_coordinate, list(name_targets), formula_sheet.title, label
            )
            if _is_pure_arithmetic(tokens):
                rewritten_tokens = Tokenizer("=" + derivation).items
                _cross_check_arithmetic(rewritten_tokens, named_values, float(cached), label)
        elif note is None:
            raise ValidationError(
                f"row {row_index}: expected value {label!r} is a bare constant without a note explaining "
                "where it comes from (rule 5)."
            )
        if any(token in label for token in REVENUE_LABEL_TOKENS) and float(cached) > 1e-9:
            raise ValidationError(
                f"row {row_index}: {label!r} is a revenue-type timeline quantity and must be negative or "
                f"zero (cost positive, revenue negative, §3.2), got {cached!r}."
            )
        example.expected.append(
            ExpectedRow(
                label=label,
                value=float(cached),
                abs_tol=float(tolerance_raw),
                derivation=derivation,
                note=note,
            )
        )

    if not example.inputs:
        raise ValidationError(f"{source_name}: the INPUTS table is empty.")
    if not example.expected:
        raise ValidationError(f"{source_name}: the EXPECTED table is empty.")
    return example, emit_yaml(example)


def emit_yaml(example: Example) -> str:
    """Renders the deterministic YAML of one example (§3.4), review block last (§3.8).

    Rendering is done by hand rather than through `yaml.dump` so that key order, float formatting
    and quoting are fully under this file's control: the CI drift check compares the regenerated
    text byte for byte, so a PyYAML version bump changing its emitter style would otherwise look
    like every example drifted at once. The result is parsed back with `yaml.safe_load` as a final
    sanity check before it is handed to the caller.

    The review block is appended only when the fingerprint of the just-rendered content matches the
    one attested in the workbook — that is the entire auto-reset mechanism of §3.8. Because it is
    written last, `content_for_fingerprint` can strip it again by a simple prefix cut.

    Returns:
        The complete YAML text, ending in a newline.

    Raises:
        ValidationError: If the rendered text does not parse back into a mapping with an `expected`
            key (a guard against a malformed cell breaking the document structure).
    """
    metadata = example.metadata
    lines = [GENERATED_HEADER.format(source=example.source_name).rstrip("\n")]
    lines.append(f"name: {quote(metadata['name'])}")
    lines.append(f"group: {quote(example.group)}")
    lines.append(f"spec_section: {quote(metadata['spec_section'])}")
    lines.append(f"computed_by: {quote(metadata['computed_by'])}")
    lines.append(f"description: {quote(metadata['description'])}")
    lines.append("inputs:")
    for label, value in example.inputs:
        lines.append(f"  {label}: {format_scalar(value)}")
    lines.append("expected:")
    for row in example.expected:
        lines.append(f"  {row.label}:")
        lines.append(f"    value: {format_value(row.value, row.abs_tol)}")
        lines.append(f"    abs_tol: {format_tolerance(row.abs_tol)}")
        if row.derivation is not None:
            lines.append(f"    derivation: {quote(row.derivation)}")
        if row.note is not None:
            lines.append(f"    note: {quote(row.note)}")
    text = "\n".join(lines) + "\n"

    fingerprint = fingerprint_of(text)
    if metadata["reviewed_fingerprint"] == fingerprint:
        text += "review:\n"
        text += f"  reviewed_by: {quote(metadata['reviewed_by'])}\n"
        text += f"  review_date: {quote(metadata['review_date'])}\n"
        text += f"  fingerprint: {quote(fingerprint)}\n"
    parsed = yaml.safe_load(text)
    if not isinstance(parsed, dict) or "expected" not in parsed:
        raise ValidationError(f"{example.source_name}: generated YAML is malformed.")
    return text


def review_status(example: Example, yaml_text: str) -> Tuple[str, str]:
    """Returns (state, message) for one converted example; state is ok / stale / unreviewed.

    Classifies the §3.8 attestation of a single example and formats the line the CLI prints. The
    three states are distinguished on purpose: `stale` means someone *did* review an earlier
    revision and the content has changed since — which is a different signal to a reviewer than an
    example nobody has ever checked. In both non-ok cases the message ends in the current
    fingerprint, which is exactly the string the reviewer types back into the workbook.

    Returns:
        `("ok" | "stale" | "unreviewed", message)`.
    """
    fingerprint = fingerprint_of(content_for_fingerprint(yaml_text))
    name = example.metadata["name"]
    stored = example.metadata["reviewed_fingerprint"]
    if stored == fingerprint:
        return "ok", f"{name}: reviewed by {example.metadata['reviewed_by']} on {example.metadata['review_date']}"
    if stored:
        return (
            "stale",
            f"{name}: UNREVIEWED — stale review by {example.metadata['reviewed_by']} "
            f"(attested {stored}) — content fingerprint {fingerprint}",
        )
    return "unreviewed", f"{name}: UNREVIEWED — content fingerprint {fingerprint}"


# --------------------------------------------------------------------------- CLI


def find_workbooks(root: str) -> List[str]:
    """All example workbooks below `root`, sorted; template and lock files are skipped.

    Discovery is purely by convention — one `.xlsx` per example anywhere under the group
    directories — so adding an example means adding a file, with no registry to update and no
    chance of an example silently not being converted. Names starting with `_` (the shared
    `_template.xlsx`) and Excel's `~$` lock files are excluded, and the result is sorted so that
    the converter's output order does not depend on the filesystem.
    """
    workbooks = []
    for directory, _, file_names in os.walk(root):
        for file_name in sorted(file_names):
            if not file_name.endswith(".xlsx") or file_name.startswith(("_", "~$")):
                continue
            workbooks.append(os.path.join(directory, file_name))
    return sorted(workbooks)


def run(root: str, check: bool, list_unreviewed: bool) -> int:
    """Converts (or verifies) every workbook below `root`; returns the process exit code.

    The one driver behind all three modes of the tool: writing the YAML fixtures (the authoring
    step), verifying them without touching the tree (`--check`, the CI drift gate of §3.6), and
    reporting attestation state (`--list-unreviewed`, §3.8). Every workbook is converted in all
    three modes — the difference is only what is done with the result — and validation failures are
    collected rather than raised, so one run reports every broken example instead of stopping at
    the first.

    Args:
        root: Directory tree holding the group sub-directories with the workbooks.
        check: Compare the regenerated text against the committed YAML instead of writing it.
        list_unreviewed: Print the unreviewed/stale examples and write nothing.

    Returns:
        0 on success, 1 if no workbook was found or committed YAML drifted, 2 if any workbook
        failed validation. Non-zero is what makes the CI job fail.
    """
    workbooks = find_workbooks(root)
    if not workbooks:
        print(f"No worked-example workbooks found below {root}.")
        return 1
    failures: List[str] = []
    drifted: List[str] = []
    unreviewed: List[str] = []
    for path in workbooks:
        yaml_path = os.path.splitext(path)[0] + ".yaml"
        try:
            example, text = convert_workbook(path)
        except ValidationError as error:
            failures.append(f"{os.path.relpath(path, root)}: {error}")
            continue
        state, message = review_status(example, text)
        if state != "ok":
            unreviewed.append(message)
        if check:
            if not os.path.isfile(yaml_path):
                drifted.append(f"{os.path.relpath(yaml_path, root)}: missing (run the converter)")
            else:
                with open(yaml_path, encoding="utf-8") as handle:
                    committed = handle.read()
                if committed != text:
                    drifted.append(f"{os.path.relpath(yaml_path, root)}: differs from the workbook")
        elif not list_unreviewed:
            with open(yaml_path, "w", encoding="utf-8") as handle:
                handle.write(text)
            print(message)

    if list_unreviewed:
        for message in unreviewed:
            print(message)
        print(f"{len(unreviewed)} of {len(workbooks)} examples are unreviewed or stale.")
    if failures:
        print("\nValidation failures:", file=sys.stderr)
        for failure in failures:
            print(f"  {failure}", file=sys.stderr)
        return 2
    if drifted:
        print("\nThe committed YAML no longer matches its workbook (§3.6):", file=sys.stderr)
        for drift in drifted:
            print(f"  {drift}", file=sys.stderr)
        print("\nRun: python tools/convert_worked_examples.py", file=sys.stderr)
        return 1
    if check:
        print(f"{len(workbooks)} worked examples: YAML matches the workbooks.")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    """Command line entry point.

    Parses the three-flag interface and delegates to `run()`. The default root is
    `tests/worked_examples` relative to this file, so the tool works from any working directory —
    including the bare CI checkout that runs the drift check without HiSim installed.

    Args:
        argv: Argument list for testing; `None` reads `sys.argv`.

    Returns:
        The process exit code from `run()`.
    """
    default_root = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tests", "worked_examples")
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("root", nargs="?", default=default_root, help="worked-example directory")
    parser.add_argument("--check", action="store_true", help="verify the committed YAML instead of writing it")
    parser.add_argument("--list-unreviewed", action="store_true", help="list examples without a valid attestation")
    arguments = parser.parse_args(argv)
    return run(arguments.root, arguments.check, arguments.list_unreviewed)


if __name__ == "__main__":
    sys.exit(main())
